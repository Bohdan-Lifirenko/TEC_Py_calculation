from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

wb = load_workbook("../data/exel/cerkassy_test.xlsx")

print(wb.sheetnames)


sht_kot1 = wb["Котел I черга"]

def get_cell(sheet: Worksheet, row: int, col: int):
    cell_value = sheet.cell(row=row, column=col).value
    return 0 if cell_value is None else float(cell_value)

def set_cell(sheet: Worksheet, row: int, col: int, value):
    sheet.cell(row=row, column=col).value = value

value = get_cell(sheet=sht_kot1, row=13, col=5)
print(value)
set_cell(sheet=sht_kot1, row=14, col=5, value=4)
print(get_cell(sheet=sht_kot1, row=14, col=5))

wb.save("cerkassy_test_NEW.xlsx")

wb = load_workbook("../data/exel/cerkassy_test.xlsx", keep_vba=True)
print(get_cell(sheet=sht_kot1, row=14, col=5))


# for row in sht_kot1.iter_rows(values_only=True):
#     print(row)