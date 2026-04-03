from openpyxl import load_workbook

wb = load_workbook("G:\\other\\PTV\\Py_calculation\\data\\exel\\cerkassy_test.xlsx")

# список назв аркушів
print(wb.sheetnames)