import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

from calculations.calc_kot1 import calc_kot1s, calc_kot1m
from json_saving import save_excel_state, load_json_to_excel
from reset_computed_cells import reset_computed_kot1_cells, reset_computed_kot2_cells, \
    reset_computed_ppk_cells, reset_computed_pwk_cells, reset_calc_tep_cells, reset_computed_tur1_cells, \
    reset_computed_tur2_cells, reset_computed_zagal_cells


# ====================== ВСІ РОЗРАХУНКОВІ ФУНКЦІЇ (встав сюди свої) ======================
# ←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←
# Сюди встав весь свій старий код з функціями:
# get_cell, set_cell, calculate_tpw, process_single_column,
# calc_kot1s, calc_kot1m, reset_computed_cells
# (весь блок, який був у попередніх версіях)
# Якщо функцій немає — додай їх саме тут перед класами!
# ====================== КІНЕЦЬ РОЗРАХУНКОВИХ ФУНКЦІЙ ======================

# ====================== КЛАС ДЛЯ ІМПОРТУ .dat ======================
class ImportDatApp:
    def __init__(self, parent_frame):
        self.frame = parent_frame
        self.dat_path = tk.StringVar()
        self.xlsx_path = tk.StringVar()

        tk.Label(self.frame, text="Імпорт даних з .dat у Excel",
                 font=("Arial", 16, "bold")).pack(pady=15)

        # .dat файл
        frame_dat = tk.Frame(self.frame)
        frame_dat.pack(pady=8, padx=20, fill="x")
        tk.Label(frame_dat, text=".dat файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(frame_dat, textvariable=self.dat_path, width=50, state="readonly").pack(side="left", padx=10)
        tk.Button(frame_dat, text="Обрати .dat", command=self.select_dat).pack(side="left")

        # Excel файл
        frame_xlsx = tk.Frame(self.frame)
        frame_xlsx.pack(pady=8, padx=20, fill="x")
        tk.Label(frame_xlsx, text="Excel файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(frame_xlsx, textvariable=self.xlsx_path, width=50, state="readonly").pack(side="left", padx=10)
        tk.Button(frame_xlsx, text="Обрати .xlsx", command=self.select_xlsx).pack(side="left")

        # Кнопка імпорту
        tk.Button(self.frame, text="ВИКОНАТИ ІМПОРТ .dat → Excel",
                  font=("Arial", 12, "bold"), bg="#28a745", fg="white", height=2,
                  command=self.run_import).pack(pady=20)

        # Лог
        self.log_text = tk.Text(self.frame, height=12, width=90, state="disabled")
        self.log_text.pack(pady=10, padx=20)

    def log(self, msg):
        self.log_text.config(state="normal")
        self.log_text.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def select_dat(self):
        path = filedialog.askopenfilename(title="Виберіть .dat файл", filetypes=[("DAT файли", "*.dat")])
        if path:
            self.dat_path.set(path)
            self.log(f"Обрано .dat: {os.path.basename(path)}")

    def select_xlsx(self):
        path = filedialog.askopenfilename(title="Виберіть Excel файл", filetypes=[("Excel файли", "*.xlsx *.xlsm")])
        if path:
            self.xlsx_path.set(path)
            self.log(f"Обрано Excel: {os.path.basename(path)}")

    def run_import(self):
        dat = self.dat_path.get().strip()
        xlsx = self.xlsx_path.get().strip()
        if not dat or not xlsx:
            messagebox.showerror("Помилка", "Оберіть обидва файли!")
            return

        self.log("Початок імпорту...")
        try:
            import_dat_to_excel(dat, xlsx)          # функція нижче
            self.log("✅ Імпорт завершено успішно!")
            messagebox.showinfo("Готово", f"Дані з .dat успішно імпортовано!\nЗбережено у:\n{os.path.basename(xlsx)}")
        except Exception as e:
            self.log(f"❌ ПОМИЛКА: {e}")
            messagebox.showerror("Помилка", str(e))

# ====================== ФУНКЦІЯ ІМПОРТУ (з твого коду) ======================
def import_dat_to_excel(dat_path: str, xlsx_path: str):
    with open(dat_path, 'r', encoding='windows-1251') as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    if not lines:
        print("❌ .dat файл порожній!")
        return

    data_lines = lines[1:] if lines[0].startswith('strtDT') else lines

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    for line in data_lines:
        parts = line.split(';')
        if len(parts) < 4:
            continue

        sheet_name = parts[0].strip()
        try:
            row = int(parts[1].strip())
            col = int(parts[2].strip())
        except ValueError:
            continue

        value_str = parts[3].strip()
        if value_str == '-':
            cell_value = '-'
        else:
            try:
                cell_value = float(value_str.replace(',', '.')) if ',' in value_str else float(value_str)
            except ValueError:
                cell_value = value_str

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        ws.cell(row=row, column=col, value=cell_value)

    wb.save(xlsx_path)
    print(f"✅ Імпортовано {len(data_lines)} записів у {xlsx_path}")


# ====================== КЛАС КАЛЬКУЛЯТОРА (твій оригінал) ======================
class KotCalculatorApp:
    def __init__(self, parent_frame):
        self.frame = parent_frame
        self.file_path = tk.StringVar()

        tk.Label(self.frame, text="Десктопний калькулятор котлів", font=("Arial", 16, "bold")).pack(pady=15)

        file_frame = tk.Frame(self.frame)
        file_frame.pack(pady=10, padx=20, fill="x")
        tk.Label(file_frame, text="Excel файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(file_frame, textvariable=self.file_path, width=60, state="readonly").pack(side="left", padx=10)
        tk.Button(file_frame, text="Обрати файл", command=self.select_file, width=15).pack(side="left")

        info = tk.Label(self.frame, text="Очікувані назви листів:\nКотел I черга | Турбіна I черга | Турбіна II черга",
                        font=("Arial", 9), fg="gray", justify="center")
        info.pack(pady=5)

        self.calc_btn = tk.Button(self.frame, text="ВИКОНАТИ РОЗРАХУНОК", font=("Arial", 12, "bold"),
                                  bg="#007ACC", fg="white", height=2, command=self.run_calculation)
        self.calc_btn.pack(pady=20)

        # Після кнопки "ВИКОНАТИ РОЗРАХУНОК" додай:
        btn_frame = tk.Frame(self.frame)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="💾 Зберегти сейв (JSON)",
                  command=self.save_current_state, width=20).pack(side="left", padx=5)

        tk.Button(btn_frame, text="📂 Завантажити сейв",
                  command=self.load_save_file, width=20).pack(side="left", padx=5)

        self.status = tk.Label(self.frame, text="Готовий до роботи", fg="green", font=("Arial", 10))
        self.status.pack(pady=10)

        self.log_text = tk.Text(self.frame, height=10, width=85, state="disabled")
        self.log_text.pack(pady=10, padx=20)

    def log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel файли", "*.xlsx *.xlsm")])
        if path:
            self.file_path.set(path)
            self.log(f"Обрано файл: {os.path.basename(path)}")

    def run_calculation(self):
        if not self.file_path.get():
            messagebox.showerror("Помилка", "Оберіть Excel файл!")
            return
        # (твій код run_calculation без змін — просто замінив root на self.frame)
        self.calc_btn.config(state="disabled")
        self.status.config(text="Обчислення...", fg="orange")
        self.log("Початок розрахунку...")
        try:
            wb = openpyxl.load_workbook(self.file_path.get(), data_only=True)
            sht_kot1 = wb["Котел I черга"]
            sht_tur1 = wb["Турбіна I черга"]
            sht_tur2 = wb["Турбіна II черга"]

            self.log("Листа завантажено успішно")
            calc_kot1s(sht_kot1, sht_tur1, sht_tur2)
            self.log("Виконано CalcKot1s")
            calc_kot1m(sht_kot1, sht_tur1, sht_tur2)
            self.log("Виконано CalcKot1m")

            base, ext = os.path.splitext(self.file_path.get())
            new_path = f"{base}_РАЗРАХОВАНО_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            wb.save(new_path)

            self.log(f"Файл збережено: {os.path.basename(new_path)}")
            self.status.config(text="Розрахунок завершено успішно!", fg="green")
            messagebox.showinfo("Готово!", f"Збережено як:\n{os.path.basename(new_path)}")
        except Exception as e:
            self.log(f"ПОМИЛКА: {e}")
            messagebox.showerror("Помилка", str(e))
            self.status.config(text="Помилка", fg="red")
        finally:
            self.calc_btn.config(state="normal")

    def save_current_state(self):
        if not self.file_path.get():
            messagebox.showerror("Помилка", "Спочатку оберіть Excel файл!")
            return
        wb = openpyxl.load_workbook(self.file_path.get())
        base = os.path.splitext(self.file_path.get())[0]
        save_path = f"{base}_SEIV_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        save_excel_state(wb, save_path)
        messagebox.showinfo("Готово", f"Сейв збережено:\n{save_path}")

    def load_save_file(self):
        path = filedialog.askopenfilename(
            title="Виберіть сейв-файл",
            filetypes=[("JSON сейви", "*.json")]
        )
        if path:
            wb = openpyxl.load_workbook(self.file_path.get())
            load_json_to_excel(path, wb)
            wb.save(self.file_path.get())          # перезаписуємо поточний файл
            messagebox.showinfo("Готово", "Дані з сейву завантажено в Excel!")

# ====================== КЛАС ОБНУЛЕННЯ (твій ResetApp) ======================
class ResetApp:
    def __init__(self, parent_frame):
        self.frame = parent_frame
        self.file_path = tk.StringVar()

        # Заголовок
        tk.Label(self.frame, text="Обнулити розрахункові клітинки",
                 font=("Arial", 16, "bold"), fg="#d32f2f").pack(pady=15)

        # Вибір файлу
        file_frame = tk.Frame(self.frame)
        file_frame.pack(pady=10, padx=20, fill="x")
        tk.Label(file_frame, text="Excel файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(file_frame, textvariable=self.file_path, width=55, state="readonly").pack(side="left", padx=10)
        tk.Button(file_frame, text="Обрати файл", command=self.select_file, width=15).pack(side="left")

        # === НОВИЙ БЛОК: Вибір сторінок для обнуління ===
        tk.Label(self.frame, text="Оберіть сторінки, які потрібно обнулити:",
                 font=("Arial", 11, "bold")).pack(anchor="w", padx=20, pady=(20, 5))

        # "Вибрати всі"
        self.select_all_var = tk.IntVar(value=1)
        self.select_all_cb = tk.Checkbutton(
            self.frame,
            text="✅ Вибрати ВСІ сторінки",
            variable=self.select_all_var,
            command=self.toggle_select_all,
            font=("Arial", 10)
        )
        self.select_all_cb.pack(anchor="w", padx=40)

        # # Список доступних сторінок + чекбокси
        # self.sheet_options = [
        #     {"name": "Котел I черга",     "sheet_name": "Котел I черга",     "func": reset_computed_kot1_cells},
        #     {"name": "Котел II черга",    "sheet_name": "Котел II черга",    "func": reset_computed_kot2_cells},
        #     {"name": "ППК",               "sheet_name": "ППК",               "func": reset_computed_ppk_cells},
        #     {"name": "ПВК",               "sheet_name": "ПВК",               "func": reset_computed_pwk_cells},
        #     {"name": "ТЕП",               "sheet_name": "ТЕП",               "func": reset_calc_tep_cells},
        #     {"name": "Турбіна I черга",   "sheet_name": "Турбіна I черга",   "func": reset_computed_tur1_cells},
        #     {"name": "Турбіна II черга",  "sheet_name": "Турбіна II черга",  "func": reset_computed_tur2_cells},
        #     {"name": "Загальні",          "sheet_name": "Загальні",          "func": reset_computed_zagal_cells},
        # ]

        self.sheet_options = [
            {"name": "Турбіна I черга", "sheet_name": "Турбіна I черга", "func": reset_computed_tur1_cells},
            {"name": "Турбіна II черга", "sheet_name": "Турбіна II черга", "func": reset_computed_tur2_cells},
            {"name": "Котел I черга", "sheet_name": "Котел I черга", "func": reset_computed_kot1_cells},
            {"name": "Котел II черга", "sheet_name": "Котел II черга", "func": reset_computed_kot2_cells},
            {"name": "ПВК", "sheet_name": "ПВК", "func": reset_computed_pwk_cells},
            {"name": "ТЕП", "sheet_name": "ТЕП", "func": reset_calc_tep_cells},
            {"name": "ППК", "sheet_name": "ППК", "func": reset_computed_ppk_cells},
            {"name": "Загальні", "sheet_name": "Загальні", "func": reset_computed_zagal_cells}
        ]

        self.reset_vars = {}          # name → IntVar
        for opt in self.sheet_options:
            var = tk.IntVar(value=1)  # за замовчуванням усі вибрані
            cb = tk.Checkbutton(
                self.frame,
                text=opt["name"],
                variable=var,
                font=("Arial", 10)
            )
            cb.pack(anchor="w", padx=40)
            self.reset_vars[opt["name"]] = var

        # Кнопка запуску
        self.reset_btn = tk.Button(
            self.frame,
            text="ОБНУЛИТИ ВИБРАНІ СТОРІНКИ",
            font=("Arial", 12, "bold"),
            bg="#d32f2f",
            fg="white",
            height=2,
            command=self.run_reset
        )
        self.reset_btn.pack(pady=25)

        self.status = tk.Label(self.frame, text="Готовий", fg="green", font=("Arial", 10))
        self.status.pack(pady=5)

        self.log_text = tk.Text(self.frame, height=12, width=85, state="disabled")
        self.log_text.pack(pady=10, padx=20)

    # ==================== ДОПОМІЖНІ МЕТОДИ ====================

    def toggle_select_all(self):
        """Вмикає / вимикає всі чекбокси"""
        state = self.select_all_var.get()
        for var in self.reset_vars.values():
            var.set(state)

    def log(self, message: str):
        self.log_text.config(state="normal")
        self.log_text.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel файли", "*.xlsx *.xlsm")])
        if path:
            self.file_path.set(path)
            self.log(f"Обрано: {os.path.basename(path)}")

    def run_reset(self):
        if not self.file_path.get():
            messagebox.showerror("Помилка", "Спочатку оберіть Excel-файл!")
            return

        # Збираємо вибрані сторінки
        selected = []
        for opt in self.sheet_options:
            if self.reset_vars[opt["name"]].get() == 1:
                selected.append(opt)

        if not selected:
            messagebox.showwarning("Попередження", "Не вибрано жодної сторінки для обнуління!")
            return

        self.reset_btn.config(state="disabled")
        self.status.config(text="Обнулення в процесі...", fg="orange")
        self.log("🚀 Початок обнуління...")

        try:
            wb = openpyxl.load_workbook(self.file_path.get(), data_only=False)

            for opt in selected:
                try:
                    sheet = wb[opt["sheet_name"]]
                    opt["func"](sheet)                    # викликаємо вашу функцію
                    self.log(f"✅ Обнулено: {opt['name']}")
                except KeyError:
                    self.log(f"⚠️ Аркуш «{opt['sheet_name']}» не знайдено в файлі!")
                except Exception as e:
                    self.log(f"❌ Помилка в {opt['name']}: {e}")

            # Зберігаємо з новим ім'ям
            base, ext = os.path.splitext(self.file_path.get())
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            new_path = f"{base}_ОБНУЛЕНО_{timestamp}{ext}"

            wb.save(new_path)

            self.log("🎉 Обнулення завершено успішно!")
            self.log(f"Файл збережено: {os.path.basename(new_path)}")
            self.status.config(text="Успішно обнулено!", fg="green")

            messagebox.showinfo(
                "Готово!",
                f"Обнулено {len(selected)} сторінок\n\nЗбережено як:\n{os.path.basename(new_path)}"
            )

        except Exception as e:
            self.log(f"🔥 Критична помилка: {e}")
            messagebox.showerror("Помилка", f"Не вдалося виконати обнуління:\n{str(e)}")
            self.status.config(text="Помилка!", fg="red")

        finally:
            self.reset_btn.config(state="normal")

# class ResetApp:
#     def __init__(self, parent_frame):
#         self.frame = parent_frame
#         self.file_path = tk.StringVar()
#
#         tk.Label(self.frame, text="Обнулити розрахункові клітинки в Kot1",
#                  font=("Arial", 16, "bold"), fg="#d32f2f").pack(pady=15)
#
#         file_frame = tk.Frame(self.frame)
#         file_frame.pack(pady=10, padx=20, fill="x")
#         tk.Label(file_frame, text="Excel файл:", font=("Arial", 11)).pack(side="left")
#         tk.Entry(file_frame, textvariable=self.file_path, width=55, state="readonly").pack(side="left", padx=10)
#         tk.Button(file_frame, text="Обрати файл", command=self.select_file, width=15).pack(side="left")
#
#         tk.Label(self.frame, text="Обнулить ТІЛЬКИ колонки 5,6,7,8 + колонку 9",
#                  font=("Arial", 9), fg="gray").pack(pady=10)
#
#         self.reset_btn = tk.Button(self.frame, text="ОБНУЛИТИ РОЗРАХУНКОВІ КЛІТИНКИ",
#                                    font=("Arial", 12, "bold"), bg="#d32f2f", fg="white", height=2,
#                                    command=self.run_reset)
#         self.reset_btn.pack(pady=20)
#
#         self.status = tk.Label(self.frame, text="Готовий", fg="green", font=("Arial", 10))
#         self.status.pack(pady=5)
#
#         self.log_text = tk.Text(self.frame, height=10, width=85, state="disabled")
#         self.log_text.pack(pady=10, padx=20)
#
#     def log(self, message):
#         self.log_text.config(state="normal")
#         self.log_text.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
#         self.log_text.see("end")
#         self.log_text.config(state="disabled")
#
#     def select_file(self):
#         path = filedialog.askopenfilename(filetypes=[("Excel файли", "*.xlsx *.xlsm")])
#         if path:
#             self.file_path.set(path)
#             self.log(f"Обрано: {os.path.basename(path)}")
#
#     def run_reset(self):
#         if not self.file_path.get():
#             messagebox.showerror("Помилка", "Оберіть файл!")
#             return
#         self.reset_btn.config(state="disabled")
#         self.status.config(text="Обнулення...", fg="orange")
#         self.log("Початок обнуління...")
#         try:
#             wb = openpyxl.load_workbook(self.file_path.get())
#             sht_kot1 = wb["Котел I черга"]
#             reset_computed_cells(sht_kot1)
#
#             base, ext = os.path.splitext(self.file_path.get())
#             new_path = f"{base}_ОБНУЛЕНО_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
#             wb.save(new_path)
#
#             self.log("✅ Обнулення завершено!")
#             self.log(f"Файл збережено: {os.path.basename(new_path)}")
#             self.status.config(text="Обнулено успішно!", fg="green")
#             messagebox.showinfo("Готово!", f"Збережено як:\n{os.path.basename(new_path)}")
#         except Exception as e:
#             self.log(f"ПОМИЛКА: {e}")
#             messagebox.showerror("Помилка", str(e))
#             self.status.config(text="Помилка", fg="red")
#         finally:
#             self.reset_btn.config(state="normal")
# ====================== ГОЛОВНЕ ВІКНО З ВКЛАДКАМИ ======================
if __name__ == "__main__":
    root = tk.Tk()
    root.title("TEC — Повний інструмент (Калькулятор + Імпорт + Обнулити)")
    root.geometry("920x680")
    root.resizable(True, True)

    style = ttk.Style()
    style.theme_use('clam')

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # Вкладка 1 — Калькулятор
    tab_calc = ttk.Frame(notebook)
    notebook.add(tab_calc, text="Калькулятор котлів")
    KotCalculatorApp(tab_calc)

    # Вкладка 2 — Імпорт .dat
    tab_import = ttk.Frame(notebook)
    notebook.add(tab_import, text="Імпорт .dat → Excel")
    ImportDatApp(tab_import)

    # Вкладка 3 — Обнулити клітинки
    tab_reset = ttk.Frame(notebook)
    notebook.add(tab_reset, text="Обнулити клітинки Kot1")
    ResetApp(tab_reset)

    root.mainloop()