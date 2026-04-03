import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import os
from datetime import datetime


def set_cell(sheet: Worksheet, row: int, col: int, value):
    """Просто записує значення в клітинку"""
    sheet.cell(row=row, column=col).value = value


def reset_computed_kot1_cells(sht_kot1: Worksheet):
    """Обнуляє ТІЛЬКИ клітинки, які розраховує твій код"""

    # ================== КОЛОНКИ 5, 6, 7, 8 ==================
    computed_cols = [5, 6, 7, 8]

    # Список всіх рядків з ProcessColumn + Process21Branch + Process22Branch
    process_column_rows = [14, 15, 19, 20, 22, 21, 23, 28, 42, 43, 44, 52, 59, 62, 76, 82, 89, 94,
                           104, 106, 108, 109, 110, 111, 112, 124, 125, 129, 130, 137, 138]

    process21_rows = [34, 37, 38, 46, 48, 53, 55, 60, 63, 65, 70, 72, 73, 74, 90, 92, 93, 80, 83, 85,
                      95, 97, 98, 100, 101, 103, 118, 120, 126, 128, 131, 133]

    process22_rows = [36, 39, 40, 49, 51, 56, 58, 61, 66, 68, 71, 73, 75, 81, 86, 88, 95, 97,
                      121, 123, 134, 136]

    all_rows_for_5_8 = list(set(process_column_rows + process21_rows + process22_rows))

    # Обнуляємо всі перелічені рядки в колонках 5-8
    for col in computed_cols:
        for row in all_rows_for_5_8:
            set_cell(sht_kot1, row, col, 0)

    # ================== КОЛОНКА 9 ==================
    column9_rows = [12, 13, 14, 15, 16, 17, 21, 22, 77, 78, 79, 110, 111, 112,
                    113, 114, 115, 116, 125, 129, 130, 138]

    for row in column9_rows:
        set_cell(sht_kot1, row, 9, 0)

    print("✅ Успішно обнулено всі розрахункові клітинки")

def reset_computed_kot2_cells(sht_kot2: Worksheet):
    """Обнуляє ТІЛЬКИ клітинки, які розраховує VBA код Kot2"""

    # ================== КОЛОНКИ 5–9 ==================
    computed_cols = [5, 6, 7, 8, 9]

    # --- БАЗОВІ + ЗАГАЛЬНІ + ВУГІЛЛЯ + ГАЗ ---
    base_rows = [14, 15, 19, 20, 22, 21, 23, 25]

    coal_rows = [28, 33, 36, 37, 42, 44, 54, 56, 61, 63, 68, 71, 73,
                 78, 80, 81, 97, 99, 100, 87, 90, 92, 105, 107, 108,
                 110, 124, 126, 132, 134, 139, 141]

    gas_rows = [29, 35, 38, 39, 45, 47, 57, 59, 64, 66, 69, 74, 76,
                79, 80, 82, 88, 93, 95, 127, 129, 142, 144]

    common_rows = [48, 50, 51, 52, 60, 67, 70, 83, 89, 96, 101,
                   102, 104, 111, 112, 114, 115, 116, 117, 118,
                   130, 131, 135, 136, 137, 138, 145, 146]

    all_rows_5_9 = list(set(base_rows + coal_rows + gas_rows + common_rows))

    # Обнуляємо колонки 5–9
    for col in computed_cols:
        for row in all_rows_5_9:
            set_cell(sht_kot2, row, col, 0)

    # ================== КОЛОНКА 10 ==================
    column10_rows = [
        12, 13, 14, 15, 16, 17,
        21, 22,
        84, 85, 86,
        116, 117, 118,
        119, 120, 121, 122,
        131, 135, 136, 146
    ]

    for row in column10_rows:
        set_cell(sht_kot2, row, 10, 0)

    print("✅ Успішно обнулено всі розрахункові клітинки Kot2")

def reset_computed_ppk_cells(sht_ppk: Worksheet):
    """
    Обнулення всіх розрахункових клітинок (аналог VBA логіки)
    ws — worksheet (наприклад wb["PPK"])
    """

    # 🔹 1. Колонки ВК (5, 6, 7)
    calc_rows_vk = [
        6, 10, 12, 14, 15, 17, 21,
        22, 23, 24, 26, 27,
        28, 30, 31,
        32, 34, 35, 37,
        38, 39, 40, 41,
        42, 44, 45
    ]

    for col in [5, 6, 7]:
        for row in calc_rows_vk:
            sht_ppk.cell(row=row, column=col).value = 0

    # 🔹 2. Підсумкова колонка (8)
    calc_rows_col8 = [
        6, 7, 11, 13, 40, 41, 45,
        48, 50, 52,
        *range(58, 66),  # 58–65
        66,
        68, 69,
        70, 72, 73,
        74, 76, 77, 78, 80, 81, 83,
        84, 86, 87, 88,
        89, 90, 91, 93, 94
    ]

    for row in calc_rows_col8:
        sht_ppk.cell(row=row, column=8).value = 0

from openpyxl.worksheet.worksheet import Worksheet

def reset_computed_pwk_cells(sht_pwk: Worksheet):
    """Обнуляє ТІЛЬКИ клітинки, які розраховує PWK VBA код"""

    # ================== КОЛОНКИ 5–9 ==================
    computed_cols = [5, 6, 7, 8, 9]

    # Всі розрахункові рядки з ProcessPWKColumn
    process_rows = [
        7, 15, 16, 17, 20, 21, 43,

        # Вугілля
        27, 30, 32, 37, 39, 44, 46, 53,
        56, 58, 75, 77, 83, 85,

        # Газ
        28, 33, 35, 40, 42, 45, 47, 49,
        54, 59, 61, 78, 80,

        # Зважені
        36, 48, 55, 62, 81,

        # Загальні
        63, 65, 66, 67, 68, 69, 70, 82,

        # Фінальні (дораховуються в CalcPWKm)
        86, 88, 89, 90, 91
    ]

    # Обнуляємо всі клітинки для колонок 5–9
    for col in computed_cols:
        for row in process_rows:
            sht_pwk.cell(row=row, column=col).value = 0

    # ================== КОЛОНКА 10 ==================

    column10_rows = [
        # коефіцієнти
        50, 51, 52,

        # підсумки
        6,

        # агрегати
        68, 69, 70,

        # середньозважені
        71, 72, 73,

        # суми
        82,
        90,
        91
    ]

    for row in column10_rows:
        sht_pwk.cell(row=row, column=10).value = 0

    print("✅ Успішно обнулено всі розрахункові клітинки PWK")

def reset_calc_tep_cells(sht_tep: Worksheet):
    """Обнуляє всі клітинки (row, 5), які розраховує CalcTEPs"""

    col = 5

    # ================== БАЗОВІ ==================
    base_rows = [8, 10, 13, 14, 16, 18]

    # ================== ККД ==================
    eff_rows = [21]

    # ================== 24–70 ==================
    main_rows = [
        24, 25, 27, 29, 31, 33, 35, 36, 37, 39,
        40, 41, 43, 45, 46, 48, 49, 51, 52, 54,
        55, 57, 58, 60, 61, 62, 64, 65, 67, 68,
        69, 70
    ]

    # ================== 71–76 ==================
    loss_rows = [71, 72, 74, 75, 76]

    # ================== 77–84 ==================
    heat_loss_rows = [78, 77, 79, 80, 81, 82, 84]

    # ================== 96–116 ==================
    energy_rows = [
        96, 97, 99, 100, 103, 104, 106,
        112, 113, 115, 116
    ]

    # ================== 117–126 ==================
    aggregate_rows = [
        117, 119, 120, 123, 124,
        107, 108, 109, 110,
        125, 126
    ]

    # ================== 128–148 ==================
    final_rows_1 = [
        128, 129, 131, 132, 134,
        135, 137, 138, 140, 141,
        143, 144, 145, 146, 148
    ]

    # ================== 152–170 ==================
    final_rows_2 = [
        152, 153, 154, 155, 156, 157,
        158, 159, 160, 161, 163,
        165, 164, 166, 167, 168,
        169, 170
    ]

    # Об'єднуємо всі
    all_rows = list(set(
        base_rows + eff_rows + main_rows +
        loss_rows + heat_loss_rows +
        energy_rows + aggregate_rows +
        final_rows_1 + final_rows_2
    ))

    # Обнулення
    for row in all_rows:
        set_cell(sht_tep, row, col, 0)

    print(f"✅ Обнулено {len(all_rows)} клітинок CalcTEPs")

def reset_computed_tur1_cells(sht_tur1: Worksheet):
    """Обнуляє всі клітинки, які розраховує CalcTur1s"""

    # ================== КОЛОНКИ 5 І 6 ==================
    computed_cols = [5, 6]

    # --- Специфічні витрати ---
    spec_rows = [12, 15, 18, 21, 24, 27, 30, 33]

    # --- Базові + підсумки ---
    base_rows = [7, 8, 9]

    # --- Середні коефіцієнти ---
    avg_rows = [35, 41]

    # --- PKNF ---
    pknf_rows = [46, 47]

    # --- Базові параметри ---
    base_param_rows = [55, 57, 58, 59, 61, 62]

    # --- Оптимальні коефіцієнти ---
    opt_rows = [65, 66, 67, 68]

    # --- Ітераційний блок ---
    iter_rows = [139]

    # --- Великі блоки (з коду) ---
    extended_rows = (
        list(range(81, 95)) +
        list(range(102, 110)) +
        list(range(115, 123)) +
        list(range(140, 158))
    )

    # Об'єднуємо
    all_rows_5_6 = list(set(
        spec_rows + base_rows + avg_rows +
        pknf_rows + base_param_rows +
        opt_rows + iter_rows + extended_rows
    ))

    # Обнуляємо колонки 5 і 6
    for col in computed_cols:
        for row in all_rows_5_6:
            set_cell(sht_tur1, row, col, 0)

    # ================== КОЛОНКА 7 (АГРЕГАЦІЯ) ==================

    col7_rows = [
        # підсумки
        7, 10, 13, 16, 19, 22, 25, 28, 31,
        57, 58, 59,

        # середні
        130, 135,

        # далі
        136, 138, 146,
        147, 149,
        150,
        157
    ]

    for row in col7_rows:
        set_cell(sht_tur1, row, 7, 0)

    print(f"✅ Tur1: обнулено {len(all_rows_5_6)} рядків у колонках 5-6 + колонка 7")

def reset_computed_tur2_cells(sht_tur2: Worksheet):
    """Обнуляє всі клітинки, які розраховує CalcTur2"""

    # ================== КОЛОНКИ 5, 6, 7 ==================
    computed_cols = [5, 6, 7]

    # --- Специфічні витрати ---
    spec_rows = [12, 15, 18, 21, 24, 27, 30, 33]

    # --- Базові ---
    base_rows = [7, 8, 9]

    # --- Середні коефіцієнти ---
    avg_rows = [35, 43]

    # --- PKNF ---
    pknf_rows = [48, 49]

    # --- Базові параметри ---
    base_param_rows = [57, 59, 60, 61, 63, 64]

    # --- Оптимальні коефіцієнти ---
    opt_rows = [67, 68, 69, 70]

    # --- Ентальпії ---
    enthalpy_rows = [
        71, 72, 73, 74,
        79, 80, 81, 82
    ]

    # --- Суха пара ---
    steam_rows = [87, 88, 89, 90, 91, 92, 93]

    # --- Великі блоки ---
    extended_rows = (
        list(range(94, 145)) +   # 94–144
        list(range(152, 168))    # 152–167
    )

    # --- Ітераційний блок ---
    iter_rows = [148]

    # --- Фінальні ---
    final_rows = [149, 150, 151]

    # Об'єднуємо всі рядки для колонок 5–7
    all_rows_5_7 = list(set(
        spec_rows + base_rows + avg_rows +
        pknf_rows + base_param_rows +
        opt_rows + enthalpy_rows +
        steam_rows + extended_rows +
        iter_rows + final_rows
    ))

    # Обнуляємо
    for col in computed_cols:
        for row in all_rows_5_7:
            set_cell(sht_tur2, row, col, 0)

    # ================== КОЛОНКА 8 (АГРЕГАЦІЯ) ==================

    col8_rows = [
        # підсумки
        7, 10, 13, 16, 19, 22, 25, 28, 31,
        59, 60, 61,

        # усереднення
        139, 144,

        # далі
        145, 147, 156,

        # коефіцієнти
        157, 159,

        # фінал
        160,
        167
    ]

    for row in col8_rows:
        set_cell(sht_tur2, row, 8, 0)

    print(f"✅ Tur2: обнулено {len(all_rows_5_7)} рядків у колонках 5-7 + колонка 8")

def reset_computed_zagal_cells(sheet: Worksheet):
    """Обнуляє всі розрахункові клітинки (CalcZagals)"""

    computed_cells = [
        (8,6),(9,6),(50,6),(51,6),(12,6),(13,6),(14,6),
        (16,6),(18,6),(20,6),(21,6),(22,6),(23,6),(24,6),(26,6),(25,6),
        (27,6),(29,6),(30,6),(28,6),
        (32,6),(33,6),(34,6),(35,6),(31,6),
        (36,6),(37,6),(38,6),
        (41,6),(42,6),(43,6),(40,6),(39,6),
        (44,6),(47,6),(48,6),(49,6),(46,6),(45,6),
    ]

    for row, col in computed_cells:
        set_cell(sheet, row, col, 0)

    print("✅ CalcZagals клітинки обнулено")