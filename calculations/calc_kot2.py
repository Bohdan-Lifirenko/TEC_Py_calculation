import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# =============================================
# ДОПОМІЖНІ ФУНКЦІЇ
# =============================================

def calculate_tpw_ws(sht_tur1: Worksheet, sht_tur2: Worksheet):
    """Розрахунок TPW1–TPW5 (однаковий для всіх стовпців)"""
    tpw1 = 0.0
    if sht_tur1.cell(row=8, column=5).value != 0:
        val = sht_tur1.cell(row=139, column=5).value or 0.0
        tpw1 = 154.7 + 0.53667 * val - 0.00088 * val ** 2

    tpw2 = 0.0
    if sht_tur1.cell(row=8, column=6).value != 0:
        val = sht_tur1.cell(row=139, column=6).value or 0.0
        tpw2 = 156.6 + 0.52364 * val - 0.00083 * val ** 2

    tpw3 = 0.0
    if sht_tur2.cell(row=8, column=5).value != 0:
        val = sht_tur2.cell(row=148, column=5).value or 0.0
        tpw3 = 184 + 0.201104 * val - 0.0001689 * val ** 2

    tpw4 = 0.0
    if sht_tur2.cell(row=8, column=6).value != 0:
        val = sht_tur2.cell(row=148, column=6).value or 0.0
        tpw4 = 184 + 0.201104 * val - 0.0001689 * val ** 2

    tpw5 = 0.0
    if sht_tur2.cell(row=8, column=7).value != 0:
        val = sht_tur2.cell(row=148, column=7).value or 0.0
        tpw5 = 184 + 0.201104 * val - 0.0001689 * val ** 2

    return tpw1, tpw2, tpw3, tpw4, tpw5


def clear_column_data(sht_kot2: Worksheet, col: int):
    """Обнулення всіх розрахункових рядків у стовпці"""
    for k in range(12, 26):
        sht_kot2.cell(row=k, column=col).value = 0
    for k in range(28, 53):
        sht_kot2.cell(row=k, column=col).value = 0
    for k in range(54, 77):
        sht_kot2.cell(row=k, column=col).value = 0
    for k in range(78, 84):
        sht_kot2.cell(row=k, column=col).value = 0
    for k in range(87, 119):
        sht_kot2.cell(row=k, column=col).value = 0
    for k in range(124, 147):
        sht_kot2.cell(row=k, column=col).value = 0


def set_coal_zero(sht_kot2: Worksheet, col: int):
    """Обнулення при shareCoal = 0"""
    rows = [28, 30, 32, 36, 38, 42, 44, 54, 56, 61, 63, 68, 71, 73, 78, 80, 81,
            87, 90, 92, 97, 99, 100, 105, 107, 108, 110, 124, 126, 132, 134,
            137, 139, 141]
    for r in rows:
        sht_kot2.cell(row=r, column=col).value = 0


def set_gas_zero(sht_kot2: Worksheet, col: int):
    """Обнулення при shareGas = 0"""
    rows = [29, 33, 35, 39, 41, 45, 47, 57, 59, 64, 66, 69, 74, 76, 79, 80, 82,
            88, 93, 95, 127, 129, 138, 142, 144]
    for r in rows:
        sht_kot2.cell(row=r, column=col).value = 0


def process_column(sht_kot2: Worksheet, sht_tur1: Worksheet, sht_tur2: Worksheet,
                   col: int, tpw1: float, tpw2: float, tpw3: float, tpw4: float, tpw5: float):
    """ОСНОВНА ЛОГІКА ДЛЯ ОДНОГО СТОВПЦЯ (5,6,7,8,9)"""
    if sht_kot2.cell(row=17, column=col).value == 0:
        clear_column_data(sht_kot2, col)
        return

    # --- Базові розрахунки ---
    sht_kot2.cell(row=14, column=col).value = (
        (sht_kot2.cell(row=12, column=col).value or 0) +
        (sht_kot2.cell(row=13, column=col).value or 0)
    )
    sht_kot2.cell(row=15, column=col).value = (
        (sht_kot2.cell(row=16, column=col).value or 0) /
        (sht_kot2.cell(row=17, column=col).value or 1)
    )
    sht_kot2.cell(row=19, column=col).value = 189500
    sht_kot2.cell(row=20, column=col).value = (
        (sht_kot2.cell(row=18, column=col).value or 0) - 189500
    )
    sht_kot2.cell(row=22, column=col).value = (
        (sht_kot2.cell(row=13, column=col).value or 0) /
        (sht_kot2.cell(row=14, column=col).value or 1)
    )
    sht_kot2.cell(row=21, column=col).value = 1 - (sht_kot2.cell(row=22, column=col).value or 0)

    # Зважена TPW
    denom = (
        (sht_tur1.cell(row=139, column=5).value or 0) +
        (sht_tur1.cell(row=139, column=6).value or 0) +
        (sht_tur2.cell(row=148, column=5).value or 0) +
        (sht_tur2.cell(row=148, column=6).value or 0) +
        (sht_tur2.cell(row=148, column=7).value or 0)
    )
    if denom == 0:
        weighted_tpw = 0.0
    else:
        weighted_tpw = (
            tpw1 * (sht_tur1.cell(row=139, column=5).value or 0) +
            tpw2 * (sht_tur1.cell(row=139, column=6).value or 0) +
            tpw3 * (sht_tur2.cell(row=148, column=5).value or 0) +
            tpw4 * (sht_tur2.cell(row=148, column=6).value or 0) +
            tpw5 * (sht_tur2.cell(row=148, column=7).value or 0)
        ) / denom
    sht_kot2.cell(row=23, column=col).value = weighted_tpw
    sht_kot2.cell(row=25, column=col).value = (
        (sht_kot2.cell(row=23, column=col).value or 0) +
        (sht_kot2.cell(row=24, column=col).value or 0)
    )

    load_factor = sht_kot2.cell(row=15, column=col).value or 0.0
    share_coal = sht_kot2.cell(row=21, column=col).value or 0.0
    share_gas = sht_kot2.cell(row=22, column=col).value or 0.0

    # --- Вугільна гілка ---
    if share_coal == 0:
        set_coal_zero(sht_kot2, col)
    else:
        lf = load_factor
        sht_kot2.cell(row=28, column=col).value = (
            (sht_kot2.cell(row=27, column=10).value or 0) +
            (4.6706 + 0.0209 * lf + 0.0000737819 * lf ** 2)
        )
        sht_kot2.cell(row=33, column=col).value = 60
        sht_kot2.cell(row=36, column=col).value = (
            (sht_kot2.cell(row=32, column=col).value or 0) -
            ((sht_kot2.cell(row=30, column=col).value or 0) +
             (4.6706 + 0.0209 * lf + 0.0000737819 * lf ** 2))
        )
        sht_kot2.cell(row=37, column=col).value = (
            60 - (26 + (4.6706 + 0.0209 * lf + 0.0000737819 * lf ** 2))
        )
        sht_kot2.cell(row=42, column=col).value = 0.988
        sht_kot2.cell(row=44, column=col).value = (
            sht_kot2.cell(row=42, column=col).value *
            (1 + (sht_kot2.cell(row=43, column=col).value or 0) / 100)
        )

        sht_kot2.cell(row=54, column=col).value = (
            1.3104 - 0.000698889 * lf + 0.00000274626 * lf ** 2
        )
        sht_kot2.cell(row=56, column=col).value = (
            sht_kot2.cell(row=54, column=col).value *
            (1 + (sht_kot2.cell(row=55, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=61, column=col).value = (
            0.1831 - 0.000857634 * lf + 0.00000166374 * lf ** 2
        )
        sht_kot2.cell(row=63, column=col).value = (
            sht_kot2.cell(row=61, column=col).value *
            (1 + (sht_kot2.cell(row=62, column=col).value or 0) / 100)
        )

        sht_kot2.cell(row=68, column=col).value = (
            (sht_kot2.cell(row=56, column=col).value or 0) +
            (sht_kot2.cell(row=63, column=col).value or 0)
        )
        sht_kot2.cell(row=71, column=col).value = 112.4231 + 0.298 * lf
        sht_kot2.cell(row=73, column=col).value = (
            (sht_kot2.cell(row=71, column=col).value or 0) +
            (sht_kot2.cell(row=72, column=col).value or 0)
        )
        sht_kot2.cell(row=78, column=col).value = (
            ((sht_kot2.cell(row=32, column=col).value or 0) - 60) * 0.6
        )
        sht_kot2.cell(row=80, column=col).value = (
            ((sht_kot2.cell(row=25, column=col).value or 0) - 226) * 0.2
        )
        sht_kot2.cell(row=81, column=col).value = (
            (sht_kot2.cell(row=73, column=col).value or 0) +
            (sht_kot2.cell(row=78, column=col).value or 0) +
            (sht_kot2.cell(row=80, column=col).value or 0)
        )

        sht_kot2.cell(row=97, column=col).value = (
            127.5691 - 1.9149 * lf + 0.00808727 * lf ** 2
        )
        sht_kot2.cell(row=99, column=col).value = (
            sht_kot2.cell(row=97, column=col).value *
            (1 + (sht_kot2.cell(row=98, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=100, column=col).value = (
            0.2 * (sht_kot2.cell(row=99, column=col).value or 0) / (100 - (sht_kot2.cell(row=99, column=col).value or 0)) *
            7800 * (sht_kot2.cell(row=8, column=10).value or 0) /
            (sht_kot2.cell(row=7, column=10).value or 1) /
            (sht_kot2.cell(row=44, column=col).value or 1)
        )

        # Складна формула 87
        part_a = (
            (sht_kot2.cell(row=84, column=10).value or 0) *
            (sht_kot2.cell(row=68, column=col).value or 0) +
            (sht_kot2.cell(row=85, column=10).value or 0)
        )
        part_b = (
            (sht_kot2.cell(row=81, column=col).value or 0) -
            (sht_kot2.cell(row=68, column=col).value or 0) *
            (sht_kot2.cell(row=28, column=col).value or 0) /
            ((sht_kot2.cell(row=68, column=col).value or 0) +
             (sht_kot2.cell(row=86, column=10).value or 0))
        )
        part_c = 0.9805 + 1.3 * (sht_kot2.cell(row=81, column=col).value or 0) / 10000
        part_d = 1 - 0.01 * (sht_kot2.cell(row=100, column=col).value or 0)
        sht_kot2.cell(row=87, column=col).value = (
            part_a * (sht_kot2.cell(row=44, column=col).value or 0) * part_b * part_c * part_d / 100 +
            0.2 * 0.15 * (sht_kot2.cell(row=8, column=10).value or 0) *
            (sht_kot2.cell(row=81, column=col).value or 0) /
            (sht_kot2.cell(row=7, column=10).value or 1)
        )

        sht_kot2.cell(row=90, column=col).value = -0.1823 + 0.00271 * lf
        sht_kot2.cell(row=92, column=col).value = (
            sht_kot2.cell(row=90, column=col).value *
            (1 + (sht_kot2.cell(row=91, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=105, column=col).value = (
            0.8 * 0.27 * 1400 * (sht_kot2.cell(row=8, column=10).value or 0) *
            (sht_kot2.cell(row=44, column=col).value or 0) /
            (sht_kot2.cell(row=7, column=10).value or 1)
        )
        sht_kot2.cell(row=107, column=col).value = (
            (sht_kot2.cell(row=105, column=col).value or 0) *
            (1 + (sht_kot2.cell(row=106, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=108, column=col).value = 0.9
        sht_kot2.cell(row=110, column=col).value = (
            sht_kot2.cell(row=108, column=col).value *
            (1 + (sht_kot2.cell(row=109, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=124, column=col).value = (
            26.8733 - 0.0873 * lf + 0.00000693577 * lf ** 2
        )
        sht_kot2.cell(row=126, column=col).value = (
            sht_kot2.cell(row=124, column=col).value *
            (1 + (sht_kot2.cell(row=125, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=132, column=col).value = 19.09
        sht_kot2.cell(row=134, column=col).value = (
            sht_kot2.cell(row=132, column=col).value *
            (1 + (sht_kot2.cell(row=133, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=139, column=col).value = (
            0.5046 + 0.0123 * lf + 0.0000128603 * lf ** 2
        )
        sht_kot2.cell(row=141, column=col).value = (
            sht_kot2.cell(row=139, column=col).value *
            (1 + (sht_kot2.cell(row=140, column=col).value or 0) / 100)
        )

    # --- Газова гілка ---
    if share_gas == 0:
        set_gas_zero(sht_kot2, col)
    else:
        lf = load_factor
        sht_kot2.cell(row=29, column=col).value = (
            (sht_kot2.cell(row=27, column=10).value or 0) +
            (-2.28 + 0.0877 * lf)
        )
        sht_kot2.cell(row=35, column=col).value = 30
        sht_kot2.cell(row=38, column=col).value = (
            (sht_kot2.cell(row=34, column=col).value or 0) -
            ((sht_kot2.cell(row=31, column=col).value or 0) - 2.28 + 0.0877 * lf)
        )
        sht_kot2.cell(row=39, column=col).value = (
            60 - (26 - 2.28 + 0.0877 * lf)
        )
        sht_kot2.cell(row=45, column=col).value = 1
        sht_kot2.cell(row=47, column=col).value = (
            sht_kot2.cell(row=45, column=col).value *
            (1 + (sht_kot2.cell(row=46, column=col).value or 0) / 100)
        )

        sht_kot2.cell(row=57, column=col).value = (
            1.8093 - 0.00749175 * lf + 0.0000206089 * lf ** 2
        )
        sht_kot2.cell(row=59, column=col).value = (
            sht_kot2.cell(row=57, column=col).value *
            (1 + (sht_kot2.cell(row=58, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=64, column=col).value = (
            0.2528 - 0.00254337 * lf + 0.0000150095 * lf ** 2 - 0.0000000345953 * lf ** 3
        )
        sht_kot2.cell(row=66, column=col).value = (
            sht_kot2.cell(row=64, column=col).value *
            (1 + (sht_kot2.cell(row=65, column=col).value or 0) / 100)
        )

        sht_kot2.cell(row=69, column=col).value = (
            (sht_kot2.cell(row=59, column=col).value or 0) +
            (sht_kot2.cell(row=66, column=col).value or 0)
        )
        sht_kot2.cell(row=74, column=col).value = (
            117.9208 + 0.04 * lf + 0.000638945 * lf ** 2
        )
        sht_kot2.cell(row=76, column=col).value = (
            (sht_kot2.cell(row=74, column=col).value or 0) +
            (sht_kot2.cell(row=75, column=col).value or 0)
        )
        sht_kot2.cell(row=79, column=col).value = (
            ((sht_kot2.cell(row=34, column=col).value or 0) - 30) * 0.6
        )
        sht_kot2.cell(row=80, column=col).value = (
            ((sht_kot2.cell(row=25, column=col).value or 0) - 226) * 0.2
        )
        sht_kot2.cell(row=82, column=col).value = (
            (sht_kot2.cell(row=76, column=col).value or 0) +
            (sht_kot2.cell(row=79, column=col).value or 0) +
            (sht_kot2.cell(row=80, column=col).value or 0)
        )

        # Складна формула 88
        part_e = 3.53 * (sht_kot2.cell(row=69, column=col).value or 0) + 0.6
        part_f = (
            (sht_kot2.cell(row=82, column=col).value or 0) -
            (sht_kot2.cell(row=69, column=col).value or 0) *
            (sht_kot2.cell(row=28, column=col).value or 0) /
            ((sht_kot2.cell(row=69, column=col).value or 0) + 0.18)
        )
        part_g = 0.9805 + 1.3 * (sht_kot2.cell(row=82, column=col).value or 0) / 10000
        sht_kot2.cell(row=88, column=col).value = part_e * part_f * part_g / 100

        sht_kot2.cell(row=93, column=col).value = 0.04
        sht_kot2.cell(row=95, column=col).value = (
            sht_kot2.cell(row=93, column=col).value *
            (1 + (sht_kot2.cell(row=94, column=col).value or 0) / 100)
        )

        if load_factor <= 82:
            sht_kot2.cell(row=127, column=col).value = (
                30.457 - 0.2398 * lf + 0.000717459 * lf ** 2
            )
        else:
            sht_kot2.cell(row=127, column=col).value = (
                58.0067 - 0.9113 * lf + 0.00713288 * lf ** 2 - 0.0000203573 * lf ** 3
            )
        sht_kot2.cell(row=129, column=col).value = (
            sht_kot2.cell(row=127, column=col).value *
            (1 + (sht_kot2.cell(row=128, column=col).value or 0) / 100)
        )
        sht_kot2.cell(row=142, column=col).value = 0.2352 + 0.00157369 * lf
        sht_kot2.cell(row=144, column=col).value = (
            sht_kot2.cell(row=142, column=col).value *
            (1 + (sht_kot2.cell(row=143, column=col).value or 0) / 100)
        )

    # --- Загальні зважені показники (після обох гілок) ---
    sht_kot2.cell(row=48, column=col).value = (
        (sht_kot2.cell(row=44, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=47, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=50, column=col).value = (sht_kot2.cell(row=49, column=col).value or 0) * 32
    sht_kot2.cell(row=51, column=col).value = (sht_kot2.cell(row=49, column=col).value or 0) * 3.6
    sht_kot2.cell(row=52, column=col).value = (sht_kot2.cell(row=49, column=col).value or 0) * 17.9

    sht_kot2.cell(row=60, column=col).value = (
        (sht_kot2.cell(row=56, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=59, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=67, column=col).value = (
        (sht_kot2.cell(row=63, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=66, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=70, column=col).value = (
        (sht_kot2.cell(row=68, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=69, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=83, column=col).value = (
        (sht_kot2.cell(row=81, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=82, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=89, column=col).value = (
        (sht_kot2.cell(row=87, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=88, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=96, column=col).value = (
        (sht_kot2.cell(row=92, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=95, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=101, column=col).value = (
        (sht_kot2.cell(row=100, column=col).value or 0) * share_coal
    )

    q5ugol = 4.2426 - 0.0762 * load_factor + 0.000566381 * load_factor ** 2 - 0.00000150637 * load_factor ** 3
    q5gaz = 2.7185 - 0.0304 * load_factor + 0.0001092 * load_factor ** 2
    sht_kot2.cell(row=102, column=col).value = q5ugol * share_coal + q5gaz * share_gas
    sht_kot2.cell(row=104, column=col).value = (
        (sht_kot2.cell(row=102, column=col).value or 0) *
        (1 + (sht_kot2.cell(row=103, column=col).value or 0) / 100)
    )

    sht_kot2.cell(row=111, column=col).value = (
        ((sht_kot2.cell(row=107, column=col).value or 0) +
         (sht_kot2.cell(row=110, column=col).value or 0)) * share_coal
    )
    sht_kot2.cell(row=112, column=col).value = 0.008 * (sht_kot2.cell(row=20, column=col).value or 0) / 100000 * 100
    sht_kot2.cell(row=114, column=col).value = (
        (sht_kot2.cell(row=113, column=col).value or 0) / 100 *
        (sht_kot2.cell(row=89, column=col).value or 0)
    )

    sht_kot2.cell(row=115, column=col).value = (
        100 - (
            (sht_kot2.cell(row=89, column=col).value or 0) +
            (sht_kot2.cell(row=96, column=col).value or 0) +
            (sht_kot2.cell(row=101, column=col).value or 0) +
            (sht_kot2.cell(row=104, column=col).value or 0) +
            (sht_kot2.cell(row=111, column=col).value or 0) +
            (sht_kot2.cell(row=112, column=col).value or 0) +
            (sht_kot2.cell(row=114, column=col).value or 0)
        )
    )

    sht_kot2.cell(row=116, column=col).value = (
        (sht_kot2.cell(row=16, column=col).value or 0) * 100 / 7 /
        (sht_kot2.cell(row=115, column=col).value or 1) +
        (sht_kot2.cell(row=50, column=col).value or 0)
    )

    if sht_kot2.cell(row=7, column=10).value == 0:
        sht_kot2.cell(row=117, column=col).value = 0
    else:
        sht_kot2.cell(row=117, column=col).value = (
            (sht_kot2.cell(row=116, column=col).value or 0) *
            7000 / (sht_kot2.cell(row=7, column=10).value or 1) * share_coal
        )

    if sht_kot2.cell(row=10, column=10).value == 0:
        sht_kot2.cell(row=118, column=col).value = 0
    else:
        sht_kot2.cell(row=118, column=col).value = (
            (sht_kot2.cell(row=116, column=col).value or 0) *
            7000 / (sht_kot2.cell(row=10, column=10).value or 1) * share_gas
        )

    sht_kot2.cell(row=130, column=col).value = (
        (sht_kot2.cell(row=126, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=129, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=131, column=col).value = (
        (sht_kot2.cell(row=130, column=col).value or 0) *
        (sht_kot2.cell(row=16, column=col).value or 0) / 1000
    )
    sht_kot2.cell(row=135, column=col).value = (
        (sht_kot2.cell(row=134, column=col).value or 0) *
        (sht_kot2.cell(row=117, column=col).value or 0) / 1000
    )
    sht_kot2.cell(row=136, column=col).value = (
        (sht_kot2.cell(row=131, column=col).value or 0) +
        (sht_kot2.cell(row=135, column=col).value or 0) +
        (sht_kot2.cell(row=51, column=col).value or 0)
    )

    sht_kot2.cell(row=137, column=col).value = (
        0.35 * (sht_kot2.cell(row=117, column=col).value or 0) *
        ((sht_kot2.cell(row=56, column=col).value or 0) - 0.08) *
        (sht_kot2.cell(row=37, column=col).value or 0) *
        ((sht_kot2.cell(row=7, column=10).value or 0) - 6 * (sht_kot2.cell(row=9, column=10).value or 0)) /
        1000000
    )
    sht_kot2.cell(row=138, column=col).value = (
        0.35 * (sht_kot2.cell(row=118, column=col).value or 0) *
        ((sht_kot2.cell(row=59, column=col).value or 0) - 0.08) *
        (sht_kot2.cell(row=39, column=col).value or 0) *
        (sht_kot2.cell(row=10, column=10).value or 0) / 1000000
    )

    sht_kot2.cell(row=145, column=col).value = (
        (sht_kot2.cell(row=141, column=col).value or 0) * share_coal +
        (sht_kot2.cell(row=144, column=col).value or 0) * share_gas
    )
    sht_kot2.cell(row=146, column=col).value = (
        (sht_kot2.cell(row=145, column=col).value or 0) *
        (sht_kot2.cell(row=17, column=col).value or 0) +
        (sht_kot2.cell(row=52, column=col).value or 0) +
        (sht_kot2.cell(row=137, column=col).value or 0) +
        (sht_kot2.cell(row=138, column=col).value or 0)
    )


# =============================================
# ОСНОВНІ ФУНКЦІЇ (тепер дуже короткі)
# =============================================

def calc_kot2s(sht_kot2: Worksheet, sht_tur1: Worksheet, sht_tur2: Worksheet):
    """CalcKot2s – для стовпців 5 та 6"""
    # Коефіцієнти ефективності (тільки для цього суба)
    if sht_kot2.cell(row=7, column=10).value == 0:
        sht_kot2.cell(row=84, column=10).value = 0
        sht_kot2.cell(row=85, column=10).value = 0
        sht_kot2.cell(row=86, column=10).value = 0
    else:
        ratio = (sht_kot2.cell(row=9, column=10).value or 0) / (sht_kot2.cell(row=7, column=10).value or 1) * 1000
        sht_kot2.cell(row=84, column=10).value = 3.5 + 0.02 * ratio
        sht_kot2.cell(row=85, column=10).value = 0.4 + 0.04 * ratio
        if ratio < 2:
            sht_kot2.cell(row=86, column=10).value = 0.14
        else:
            sht_kot2.cell(row=86, column=10).value = 0.12 + 0.014 * ratio

    tpw1, tpw2, tpw3, tpw4, tpw5 = calculate_tpw_ws(sht_tur1, sht_tur2)
    process_column(sht_kot2, sht_tur1, sht_tur2, 5, tpw1, tpw2, tpw3, tpw4, tpw5)
    process_column(sht_kot2, sht_tur1, sht_tur2, 6, tpw1, tpw2, tpw3, tpw4, tpw5)


def calc_kot2m(sht_kot2: Worksheet, sht_tur1: Worksheet, sht_tur2: Worksheet):
    """CalcKot2m – для стовпця 7"""
    tpw1, tpw2, tpw3, tpw4, tpw5 = calculate_tpw_ws(sht_tur1, sht_tur2)
    process_column(sht_kot2, sht_tur1, sht_tur2, 7, tpw1, tpw2, tpw3, tpw4, tpw5)


def calc_kot2t(sht_kot2: Worksheet, sht_tur1: Worksheet, sht_tur2: Worksheet):
    """CalcKot2t – для стовпців 8, 9 + підсумки в колонці 10"""
    tpw1, tpw2, tpw3, tpw4, tpw5 = calculate_tpw_ws(sht_tur1, sht_tur2)
    process_column(sht_kot2, sht_tur1, sht_tur2, 8, tpw1, tpw2, tpw3, tpw4, tpw5)
    process_column(sht_kot2, sht_tur1, sht_tur2, 9, tpw1, tpw2, tpw3, tpw4, tpw5)

    # Підсумки в колонці 10
    sht_kot2.cell(row=116, column=10).value = (
        (sht_kot2.cell(row=116, column=5).value or 0) +
        (sht_kot2.cell(row=116, column=6).value or 0) +
        (sht_kot2.cell(row=116, column=7).value or 0) +
        (sht_kot2.cell(row=116, column=8).value or 0) +
        (sht_kot2.cell(row=116, column=9).value or 0)
    )
    sht_kot2.cell(row=117, column=10).value = (
        (sht_kot2.cell(row=117, column=5).value or 0) +
        (sht_kot2.cell(row=117, column=6).value or 0) +
        (sht_kot2.cell(row=117, column=7).value or 0) +
        (sht_kot2.cell(row=117, column=8).value or 0) +
        (sht_kot2.cell(row=117, column=9).value or 0)
    )
    sht_kot2.cell(row=118, column=10).value = (
        (sht_kot2.cell(row=118, column=5).value or 0) +
        (sht_kot2.cell(row=118, column=6).value or 0) +
        (sht_kot2.cell(row=118, column=7).value or 0) +
        (sht_kot2.cell(row=118, column=8).value or 0) +
        (sht_kot2.cell(row=118, column=9).value or 0)
    )

    if sht_kot2.cell(row=116, column=10).value == 0:
        for r in [119, 120, 121, 122]:
            sht_kot2.cell(row=r, column=10).value = 0
        for r in [21, 22, 15]:
            sht_kot2.cell(row=r, column=10).value = 0
        for r in [131, 135, 136, 146]:
            sht_kot2.cell(row=r, column=10).value = 0
    else:
        sht_kot2.cell(row=12, column=10).value = (
            (sht_kot2.cell(row=12, column=5).value or 0) +
            (sht_kot2.cell(row=12, column=6).value or 0) +
            (sht_kot2.cell(row=12, column=7).value or 0) +
            (sht_kot2.cell(row=12, column=8).value or 0) +
            (sht_kot2.cell(row=12, column=9).value or 0)
        )
        sht_kot2.cell(row=13, column=10).value = (
            (sht_kot2.cell(row=13, column=5).value or 0) +
            (sht_kot2.cell(row=13, column=6).value or 0) +
            (sht_kot2.cell(row=13, column=7).value or 0) +
            (sht_kot2.cell(row=13, column=8).value or 0) +
            (sht_kot2.cell(row=13, column=9).value or 0)
        )
        sht_kot2.cell(row=14, column=10).value = (
            (sht_kot2.cell(row=14, column=5).value or 0) +
            (sht_kot2.cell(row=14, column=6).value or 0) +
            (sht_kot2.cell(row=14, column=7).value or 0) +
            (sht_kot2.cell(row=14, column=8).value or 0) +
            (sht_kot2.cell(row=14, column=9).value or 0)
        )
        sht_kot2.cell(row=16, column=10).value = (
            (sht_kot2.cell(row=16, column=5).value or 0) +
            (sht_kot2.cell(row=16, column=6).value or 0) +
            (sht_kot2.cell(row=16, column=7).value or 0) +
            (sht_kot2.cell(row=16, column=8).value or 0) +
            (sht_kot2.cell(row=16, column=9).value or 0)
        )

        total_17 = (
            (sht_kot2.cell(row=17, column=5).value or 0) +
            (sht_kot2.cell(row=17, column=6).value or 0) +
            (sht_kot2.cell(row=17, column=7).value or 0) +
            (sht_kot2.cell(row=17, column=8).value or 0) +
            (sht_kot2.cell(row=17, column=9).value or 0)
        )
        if total_17 == 0:
            sht_kot2.cell(row=15, column=10).value = 0
        else:
            sht_kot2.cell(row=15, column=10).value = (
                sht_kot2.cell(row=16, column=10).value / total_17
            )

        sht_kot2.cell(row=17, column=10).value = total_17

        sht_kot2.cell(row=119, column=10).value = (
            (sht_kot2.cell(row=115, column=5).value or 0) * (sht_kot2.cell(row=116, column=5).value or 0) +
            (sht_kot2.cell(row=115, column=6).value or 0) * (sht_kot2.cell(row=116, column=6).value or 0) +
            (sht_kot2.cell(row=115, column=7).value or 0) * (sht_kot2.cell(row=116, column=7).value or 0) +
            (sht_kot2.cell(row=115, column=8).value or 0) * (sht_kot2.cell(row=116, column=8).value or 0) +
            (sht_kot2.cell(row=115, column=9).value or 0) * (sht_kot2.cell(row=116, column=9).value or 0)
        ) / (sht_kot2.cell(row=116, column=10).value or 1)

        sht_kot2.cell(row=120, column=10).value = (
            (sht_kot2.cell(row=83, column=5).value or 0) * (sht_kot2.cell(row=116, column=5).value or 0) +
            (sht_kot2.cell(row=83, column=6).value or 0) * (sht_kot2.cell(row=116, column=6).value or 0) +
            (sht_kot2.cell(row=83, column=7).value or 0) * (sht_kot2.cell(row=116, column=7).value or 0) +
            (sht_kot2.cell(row=83, column=8).value or 0) * (sht_kot2.cell(row=116, column=8).value or 0) +
            (sht_kot2.cell(row=83, column=9).value or 0) * (sht_kot2.cell(row=116, column=9).value or 0)
        ) / (sht_kot2.cell(row=116, column=10).value or 1)

        sht_kot2.cell(row=121, column=10).value = (
            (sht_kot2.cell(row=67, column=5).value or 0) * (sht_kot2.cell(row=116, column=5).value or 0) +
            (sht_kot2.cell(row=67, column=6).value or 0) * (sht_kot2.cell(row=116, column=6).value or 0) +
            (sht_kot2.cell(row=67, column=7).value or 0) * (sht_kot2.cell(row=116, column=7).value or 0) +
            (sht_kot2.cell(row=67, column=8).value or 0) * (sht_kot2.cell(row=116, column=8).value or 0) +
            (sht_kot2.cell(row=67, column=9).value or 0) * (sht_kot2.cell(row=116, column=9).value or 0)
        ) / (sht_kot2.cell(row=116, column=10).value or 1)

        sht_kot2.cell(row=122, column=10).value = (
            (sht_kot2.cell(row=70, column=5).value or 0) * (sht_kot2.cell(row=116, column=5).value or 0) +
            (sht_kot2.cell(row=70, column=6).value or 0) * (sht_kot2.cell(row=116, column=6).value or 0) +
            (sht_kot2.cell(row=70, column=7).value or 0) * (sht_kot2.cell(row=116, column=7).value or 0) +
            (sht_kot2.cell(row=70, column=8).value or 0) * (sht_kot2.cell(row=116, column=8).value or 0) +
            (sht_kot2.cell(row=70, column=9).value or 0) * (sht_kot2.cell(row=116, column=9).value or 0)
        ) / (sht_kot2.cell(row=116, column=10).value or 1)

        sht_kot2.cell(row=21, column=10).value = (
            (sht_kot2.cell(row=117, column=10).value or 0) *
            (sht_kot2.cell(row=7, column=10).value or 0) /
            (sht_kot2.cell(row=116, column=10).value or 1) / 7000
        )
        sht_kot2.cell(row=22, column=10).value = (
            (sht_kot2.cell(row=118, column=10).value or 0) *
            (sht_kot2.cell(row=10, column=10).value or 0) /
            (sht_kot2.cell(row=116, column=10).value or 1) / 7000
        )

        sht_kot2.cell(row=131, column=10).value = (
            (sht_kot2.cell(row=131, column=5).value or 0) +
            (sht_kot2.cell(row=131, column=6).value or 0) +
            (sht_kot2.cell(row=131, column=7).value or 0) +
            (sht_kot2.cell(row=131, column=8).value or 0) +
            (sht_kot2.cell(row=131, column=9).value or 0)
        )
        sht_kot2.cell(row=135, column=10).value = (
            (sht_kot2.cell(row=135, column=5).value or 0) +
            (sht_kot2.cell(row=135, column=6).value or 0) +
            (sht_kot2.cell(row=135, column=7).value or 0) +
            (sht_kot2.cell(row=135, column=8).value or 0) +
            (sht_kot2.cell(row=135, column=9).value or 0)
        )
        sht_kot2.cell(row=136, column=10).value = (
            (sht_kot2.cell(row=136, column=5).value or 0) +
            (sht_kot2.cell(row=136, column=6).value or 0) +
            (sht_kot2.cell(row=136, column=7).value or 0) +
            (sht_kot2.cell(row=136, column=8).value or 0) +
            (sht_kot2.cell(row=136, column=9).value or 0)
        )
        sht_kot2.cell(row=146, column=10).value = (
            (sht_kot2.cell(row=146, column=5).value or 0) +
            (sht_kot2.cell(row=146, column=6).value or 0) +
            (sht_kot2.cell(row=146, column=7).value or 0) +
            (sht_kot2.cell(row=146, column=8).value or 0) +
            (sht_kot2.cell(row=146, column=9).value or 0)
        )


# =============================================
# ПРИКЛАД ВИКОРИСТАННЯ
# =============================================
if __name__ == "__main__":
    # Замініть на шлях до вашого Excel-файлу
    wb = openpyxl.load_workbook("/data/exel/cerkassy_test.xlsx", data_only=False)  # data_only=False щоб записувати формули/значення

    # Замініть на реальні назви аркушів (або CodeName, якщо вони відрізняються)
    sht_tur1 = wb["Турбіна I черга"]      # або wb.worksheets[індекс]
    sht_tur2 = wb["Турбіна II черга"]
    sht_kot2 = wb["Котел II черга"]

    # Викликайте в тому самому порядку, як у VBA
    calc_kot2s(sht_kot2, sht_tur1, sht_tur2)
    calc_kot2m(sht_kot2, sht_tur1, sht_tur2)
    calc_kot2t(sht_kot2, sht_tur1, sht_tur2)

    # Збережіть зміни
    wb.save("your_file_updated.xlsx")
    print("Розрахунок завершено!")