import openpyxl
from openpyxl import load_workbook

# =============================================
# ДОПОМІЖНІ ФУНКЦІЇ
# =============================================
def clear_pwk_column(ws, col: int):
    """Повне обнулення стовпця для одного ВК (коли Cells(18, col) = 0)"""
    ws.cell(row=6, column=col).value = 0
    ws.cell(row=7, column=col).value = 0
    for k in range(13, 22):
        ws.cell(row=k, column=col).value = 0
    ws.cell(row=24, column=col).value = 0
    ws.cell(row=25, column=col).value = 0
    ws.cell(row=27, column=col).value = 0
    ws.cell(row=28, column=col).value = 0
    for k in range(30, 50):
        ws.cell(row=k, column=col).value = 0
    for k in range(53, 71):
        ws.cell(row=k, column=col).value = 0
    for k in range(75, 92):
        ws.cell(row=k, column=col).value = 0


def process_pwk_column(ws, col: int):
    """Обробка одного стовпця (5–9). Логіка 100% ідентична VBA."""
    if ws.cell(row=18, column=col).value == 0 or ws.cell(row=18, column=col).value is None:
        clear_pwk_column(ws, col)
        return

    # Базові розрахунки (однакові для всіх стовпців)
    ws.cell(row=7, column=col).value = ws.cell(row=6, column=col).value / ws.cell(row=18, column=col).value
    ws.cell(row=15, column=col).value = (
        ws.cell(row=13, column=col).value * ws.cell(row=9, column=10).value +
        ws.cell(row=14, column=col).value * ws.cell(row=11, column=10).value
    ) / 7000
    ws.cell(row=16, column=col).value = (
        ws.cell(row=13, column=col).value * ws.cell(row=9, column=10).value /
        ws.cell(row=15, column=col).value / 7000
    )
    ws.cell(row=17, column=col).value = (
        ws.cell(row=14, column=col).value * ws.cell(row=11, column=10).value /
        ws.cell(row=15, column=col).value / 7000
    )

    # === Різні константи для груп стовпців ===
    if col <= 7:          # тип 1: ВК-1, ВК-2, ВК-3 (стовпці 5-7)
        row20_val = 16160
        temp_threshold = 2140
        row27_cell = 23
        row27_add = 3.0
        row28_cell = 26
        row28_add = 3.2
        row30_mult = 1.0908
        row30_sub = 1.6424
        row37_const = 88.35
        row37_mult = 0.997
        row40_const = 76.9
        row40_mult = 2.5537
        row40_quad = -0.00799556
        row33_mult = 1.1547
        row33_sub = 3.6289
        row75_mult = 2.5748
        row75_sub = 15.0159
        row78_mult = 2.6173
        row78_sub = 15.2111
        corr13_1 = 0.033
        corr13_2 = 0.02
        corr14_1 = 0.033
        corr14_2 = 0.02
        row56_val = 0.02
        use_row55_in_67 = True
    else:                 # тип 2: ВК-4, ВК-5 (стовпці 8-9)
        row20_val = 10090
        temp_threshold = 4420
        row27_cell = 26
        row27_add = 3.6
        row28_cell = 26
        row28_add = 3.8
        row30_mult = 1.0065
        row30_sub = 9.1503
        row37_const = 90.84
        row37_mult = 0.611
        row40_const = 112.39
        row40_mult = 0.591
        row40_quad = 0.0
        row33_mult = 1.0865
        row33_sub = 9.1503
        row75_mult = 4.5378
        row75_sub = 38.3363
        row78_mult = 4.5378
        row78_sub = 38.3363
        corr13_1 = 0.005
        corr13_2 = 0.0
        corr14_1 = 0.019
        corr14_2 = 0.00545
        row56_val = 0.0
        use_row55_in_67 = False

    ws.cell(row=20, column=col).value = row20_val
    ws.cell(row=21, column=col).value = ws.cell(row=19, column=col).value - row20_val

    # Коригування рядка 43
    if ws.cell(row=25, column=col).value > 104:
        ws.cell(row=43, column=col).value = 0
    else:
        ws.cell(row=43, column=col).value = 0.9 * (ws.cell(row=25, column=col).value - 104)

    # === Блок якщо Cells(13, col) = 0 ===
    if ws.cell(row=13, column=col).value == 0 or ws.cell(row=13, column=col).value is None:
        ws.cell(row=27, column=col).value = 0
        ws.cell(row=30, column=col).value = 0
        ws.cell(row=32, column=col).value = 0
        ws.cell(row=37, column=col).value = 0
        ws.cell(row=39, column=col).value = 0
        ws.cell(row=44, column=col).value = 0
        ws.cell(row=46, column=col).value = 0
        ws.cell(row=53, column=col).value = 0
        ws.cell(row=56, column=col).value = 0
        ws.cell(row=58, column=col).value = 0
        ws.cell(row=75, column=col).value = 0
        ws.cell(row=77, column=col).value = 0
        ws.cell(row=83, column=col).value = 0
        ws.cell(row=85, column=col).value = 0
    else:
        ws.cell(row=27, column=col).value = ws.cell(row=row27_cell, column=10).value + row27_add
        ws.cell(row=30, column=col).value = (
            row30_mult * ws.cell(row=7, column=col).value /
            (ws.cell(row=7, column=col).value - row30_sub)
        )
        ws.cell(row=32, column=col).value = ws.cell(row=30, column=col).value * (
            1 + ws.cell(row=31, column=col).value / 100
        )
        ws.cell(row=37, column=col).value = row37_const + row37_mult * ws.cell(row=7, column=col).value
        ws.cell(row=39, column=col).value = ws.cell(row=37, column=col).value + ws.cell(row=38, column=col).value

        if ws.cell(row=24, column=col).value > temp_threshold:
            ws.cell(row=44, column=col).value = 0
        else:
            ws.cell(row=44, column=col).value = -(
                corr13_1 + corr13_2 * ws.cell(row=7, column=col).value
            ) * (ws.cell(row=24, column=col).value - temp_threshold) / 100

        ws.cell(row=46, column=col).value = (
            ws.cell(row=39, column=col).value +
            ws.cell(row=43, column=col).value +
            ws.cell(row=44, column=col).value
        )

        # Складна формула рядка 53
        part_a = 3.53 * ws.cell(row=32, column=col).value + 0.6
        part_b = (
            ws.cell(row=46, column=col).value -
            ws.cell(row=32, column=col).value *
            ws.cell(row=27, column=col).value /
            (ws.cell(row=32, column=col).value + 0.18)
        )
        part_c = 0.9805 + 1.3 * ws.cell(row=46, column=col).value / 10000
        ws.cell(row=53, column=col).value = part_a * part_b * part_c / 100

        ws.cell(row=56, column=col).value = row56_val
        ws.cell(row=58, column=col).value = ws.cell(row=56, column=col).value * (
            1 + ws.cell(row=57, column=col).value / 100
        )
        ws.cell(row=75, column=col).value = (
            row75_mult * ws.cell(row=7, column=col).value /
            (ws.cell(row=7, column=col).value - row75_sub)
        )
        ws.cell(row=77, column=col).value = ws.cell(row=75, column=col).value * (
            1 + ws.cell(row=76, column=col).value / 100
        )
        ws.cell(row=83, column=col).value = 12
        ws.cell(row=85, column=col).value = 12 * (1 + ws.cell(row=84, column=col).value / 100)

    # === Блок якщо Cells(14, col) = 0 ===
    if ws.cell(row=14, column=col).value == 0 or ws.cell(row=14, column=col).value is None:
        ws.cell(row=28, column=col).value = 0
        ws.cell(row=33, column=col).value = 0
        ws.cell(row=35, column=col).value = 0
        ws.cell(row=40, column=col).value = 0
        ws.cell(row=42, column=col).value = 0
        ws.cell(row=45, column=col).value = 0
        ws.cell(row=47, column=col).value = 0
        ws.cell(row=49, column=col).value = 0
        ws.cell(row=54, column=col).value = 0
        ws.cell(row=59, column=col).value = 0
        ws.cell(row=61, column=col).value = 0
        ws.cell(row=78, column=col).value = 0
        ws.cell(row=80, column=col).value = 0
        ws.cell(row=86, column=col).value = 0
        ws.cell(row=88, column=col).value = 0
    else:
        ws.cell(row=28, column=col).value = ws.cell(row=row28_cell, column=10).value + row28_add
        ws.cell(row=33, column=col).value = (
            row33_mult * ws.cell(row=7, column=col).value /
            (ws.cell(row=7, column=col).value - row33_sub)
        )
        ws.cell(row=35, column=col).value = ws.cell(row=33, column=col).value * (
            1 + ws.cell(row=34, column=col).value / 100
        )

        if col <= 7:
            ws.cell(row=40, column=col).value = (
                row40_const +
                row40_mult * ws.cell(row=7, column=col).value +
                row40_quad * ws.cell(row=7, column=col).value ** 2
            )
        else:
            ws.cell(row=40, column=col).value = (
                row40_const + row40_mult * ws.cell(row=7, column=col).value
            )

        ws.cell(row=42, column=col).value = ws.cell(row=40, column=col).value + ws.cell(row=41, column=col).value

        if ws.cell(row=24, column=col).value > temp_threshold:
            ws.cell(row=45, column=col).value = 0
        else:
            ws.cell(row=45, column=col).value = -(
                corr14_1 + corr14_2 * ws.cell(row=7, column=col).value
            ) * (ws.cell(row=24, column=col).value - temp_threshold) / 100

        ws.cell(row=47, column=col).value = (
            ws.cell(row=42, column=col).value +
            ws.cell(row=43, column=col).value +
            ws.cell(row=45, column=col).value
        )
        ws.cell(row=49, column=col).value = (
            ws.cell(row=13, column=col).value + 0.993 * ws.cell(row=14, column=col).value
        )

        # Складна формула рядка 54
        part_a = (
            ws.cell(row=50, column=10).value * ws.cell(row=35, column=col).value +
            ws.cell(row=51, column=10).value
        )
        part_b = (
            ws.cell(row=47, column=col).value -
            ws.cell(row=35, column=col).value *
            ws.cell(row=28, column=col).value /
            (ws.cell(row=35, column=col).value + ws.cell(row=52, column=10).value)
        )
        part_c = 0.9805 + 1.3 * ws.cell(row=47, column=col).value / 10000
        ws.cell(row=54, column=col).value = part_a * part_b * part_c * ws.cell(row=49, column=col).value / 100

        ws.cell(row=59, column=col).value = 0
        ws.cell(row=61, column=col).value = 0
        ws.cell(row=78, column=col).value = (
            row78_mult * ws.cell(row=7, column=col).value /
            (ws.cell(row=7, column=col).value - row78_sub)
        )
        ws.cell(row=80, column=col).value = ws.cell(row=78, column=col).value * (
            1 + ws.cell(row=79, column=col).value / 100
        )

        # Попереднє значення рядка 86 (буде перезаписано в підсумках)
        ws.cell(row=86, column=col).value = (
            12 + 67.5 / ws.cell(row=14, column=10).value * ws.cell(row=14, column=col).value
        )

    # === Спільна частина після блоків 13 і 14 ===
    ws.cell(row=36, column=col).value = (
        ws.cell(row=32, column=col).value * ws.cell(row=16, column=col).value +
        ws.cell(row=35, column=col).value * ws.cell(row=17, column=col).value
    )
    ws.cell(row=48, column=col).value = (
        ws.cell(row=46, column=col).value * ws.cell(row=16, column=col).value +
        ws.cell(row=47, column=col).value * ws.cell(row=17, column=col).value
    )
    ws.cell(row=55, column=col).value = (
        ws.cell(row=53, column=col).value * ws.cell(row=16, column=col).value +
        ws.cell(row=54, column=col).value * ws.cell(row=17, column=col).value
    )
    ws.cell(row=62, column=col).value = (
        ws.cell(row=58, column=col).value * ws.cell(row=16, column=col).value +
        ws.cell(row=61, column=col).value * ws.cell(row=17, column=col).value
    )

    ws.cell(row=63, column=col).value = 0.05
    ws.cell(row=65, column=col).value = ws.cell(row=63, column=col).value * (
        1 + ws.cell(row=64, column=col).value / 100
    )
    ws.cell(row=66, column=col).value = (
        0.0025 * ws.cell(row=21, column=col).value / 100000 * 100
    )

    # Рядок 67 — різний для двох груп
    if use_row55_in_67:
        ws.cell(row=67, column=col).value = 100 - (
            ws.cell(row=55, column=col).value +
            ws.cell(row=62, column=col).value +
            ws.cell(row=65, column=col).value +
            ws.cell(row=66, column=col).value
        )
    else:
        ws.cell(row=67, column=col).value = 100 - (
            ws.cell(row=53, column=col).value +
            ws.cell(row=62, column=col).value +
            ws.cell(row=65, column=col).value +
            ws.cell(row=66, column=col).value
        )

    ws.cell(row=68, column=col).value = (
        ws.cell(row=6, column=col).value * 100 / 7 / ws.cell(row=67, column=col).value
    )

    ws.cell(row=69, column=col).value = (
        0 if ws.cell(row=13, column=col).value == 0 else
        ws.cell(row=68, column=col).value * 7000 / ws.cell(row=9, column=10).value * ws.cell(row=13, column=col).value
    )
    ws.cell(row=70, column=col).value = (
        0 if ws.cell(row=14, column=col).value == 0 else
        ws.cell(row=68, column=col).value * 7000 / ws.cell(row=11, column=10).value * ws.cell(row=14, column=col).value
    )

    ws.cell(row=81, column=col).value = (
        ws.cell(row=77, column=col).value * ws.cell(row=16, column=col).value +
        ws.cell(row=80, column=col).value * ws.cell(row=17, column=col).value
    )
    ws.cell(row=82, column=col).value = (
        ws.cell(row=81, column=col).value * ws.cell(row=6, column=col).value / 1000
    )


# =============================================
# ГОЛОВНА ФУНКЦІЯ (об’єднує CalcPWKs + CalcPWKm)
# =============================================
def calc_pwks(ws):
    """Повний розрахунок PWK (аналог VBA CalcPWKs + CalcPWKm)"""
    # === Початковий блок (з CalcPWKs) ===
    if ws.cell(row=11, column=10).value == 0 or ws.cell(row=11, column=10).value is None:
        ws.cell(row=50, column=10).value = 0
        ws.cell(row=51, column=10).value = 0
        ws.cell(row=52, column=10).value = 0
    else:
        ws.cell(row=50, column=10).value = 0.495 + 0.02 * ws.cell(row=12, column=10).value
        ws.cell(row=51, column=10).value = 0.44 + 0.04 * ws.cell(row=12, column=10).value
        ws.cell(row=52, column=10).value = 0.13

    # === Обробка всіх п'яти ВК ===
    for col in [5, 6, 7, 8, 9]:
        process_pwk_column(ws, col)

    # === Підсумковий блок (з CalcPWKm) ===
    total_18 = sum(
        ws.cell(row=18, column=c).value or 0 for c in [5, 6, 7, 8, 9]
    )

    if total_18 == 0:
        for r in [68, 69, 70, 71, 72, 73]:
            ws.cell(row=r, column=10).value = 0
    else:
        ws.cell(row=6, column=10).value = sum(
            ws.cell(row=6, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )
        ws.cell(row=68, column=10).value = sum(
            ws.cell(row=68, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )
        ws.cell(row=69, column=10).value = sum(
            ws.cell(row=69, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )
        ws.cell(row=70, column=10).value = sum(
            ws.cell(row=70, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )

        sum_68 = ws.cell(row=68, column=10).value or 0
        if sum_68 == 0:
            ws.cell(row=71, column=10).value = 0
            ws.cell(row=72, column=10).value = 0
            ws.cell(row=73, column=10).value = 0
        else:
            ws.cell(row=71, column=10).value = sum(
                (ws.cell(row=67, column=c).value or 0) * (ws.cell(row=68, column=c).value or 0)
                for c in [5, 6, 7, 8, 9]
            ) / sum_68
            ws.cell(row=72, column=10).value = sum(
                (ws.cell(row=48, column=c).value or 0) * (ws.cell(row=68, column=c).value or 0)
                for c in [5, 6, 7, 8, 9]
            ) / sum_68
            ws.cell(row=73, column=10).value = sum(
                (ws.cell(row=36, column=c).value or 0) * (ws.cell(row=68, column=c).value or 0)
                for c in [5, 6, 7, 8, 9]
            ) / sum_68

        ws.cell(row=82, column=10).value = sum(
            ws.cell(row=82, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )

        # Перерахунок рядків 86 і 88 для кожного ВК (залежно від сумарного навантаження)
        for col in [5, 6, 7, 8, 9]:
            if (
                ws.cell(row=14, column=col).value == 0 or
                ws.cell(row=18, column=col).value == 0
            ):
                ws.cell(row=86, column=col).value = 0
            else:
                ws.cell(row=86, column=col).value = (
                    12 + 67.5 / ws.cell(row=70, column=10).value * ws.cell(row=70, column=col).value
                )
                ws.cell(row=88, column=col).value = ws.cell(row=86, column=col).value * (
                    1 + ws.cell(row=87, column=col).value / 100
                )

        # Рядки 89, 90, 91
        for col in [5, 6, 7, 8, 9]:
            ws.cell(row=89, column=col).value = (
                ws.cell(row=85, column=col).value * ws.cell(row=16, column=col).value +
                ws.cell(row=88, column=col).value * ws.cell(row=17, column=col).value
            )
            ws.cell(row=90, column=col).value = (
                ws.cell(row=89, column=col).value * ws.cell(row=18, column=col).value / 1000
            )
            ws.cell(row=91, column=col).value = (
                ws.cell(row=82, column=col).value + ws.cell(row=90, column=col).value
            )

        ws.cell(row=90, column=10).value = sum(
            ws.cell(row=90, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )
        ws.cell(row=91, column=10).value = sum(
            ws.cell(row=91, column=c).value or 0 for c in [5, 6, 7, 8, 9]
        )


# =============================================
# ПРИКЛАД ВИКОРИСТАННЯ
# =============================================
if __name__ == "__main__":
    wb = load_workbook("/data/exel/cerkassy_test_ОБНУЛЕНО_20260402_144235.xlsx")          # ← шлях до твого Excel-файлу
    ws = wb["shtPWK"]                             # ← назва аркуша (CodeName або видиму назву)

    calc_pwks(ws)

    wb.save("G:\\other\\PTV\\Py_calculation\\data\\exel\\cerkassy_test_teps_pwk.xlsx")
    print("✅ Розрахунок PWK завершено! Файл збережено як your_file_updated.xlsx")