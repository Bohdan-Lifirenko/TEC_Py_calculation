import openpyxl
from openpyxl import load_workbook

# =============================================
# ДОПОМІЖНІ ФУНКЦІЇ
# =============================================
def clear_ppk_column(ws, col: int):
    """Обнулення всіх розрахункових рядків для одного ВК"""
    for k in range(6, 20):      # 6 до 19
        ws.cell(row=k, column=col).value = 0
    for k in range(21, 46):     # 21 до 45
        ws.cell(row=k, column=col).value = 0


def process_ppk_column(ws, col: int):
    """ОСНОВНА ЛОГІКА ДЛЯ ОДНОГО ВК (стовпці 5, 6 або 7)"""
    if ws.cell(row=8, column=col).value < 0.1:
        clear_ppk_column(ws, col)
        return

    # --- Базові розрахунки ---
    ws.cell(row=6, column=col).value = (
        ws.cell(row=7, column=col).value * ws.cell(row=4, column=8).value / 7000
    )
    ws.cell(row=10, column=col).value = 48650

    load_factor = ws.cell(row=11, column=col).value / ws.cell(row=8, column=col).value
    ws.cell(row=12, column=col).value = load_factor

    ws.cell(row=14, column=col).value = 618
    spec_cons = ws.cell(row=13, column=col).value / ws.cell(row=8, column=col).value
    ws.cell(row=15, column=col).value = spec_cons

    ws.cell(row=17, column=col).value = 70

    temp_before = ws.cell(row=20, column=8).value + 2
    ws.cell(row=21, column=col).value = temp_before

    # --- Коригування ---
    if ws.cell(row=19, column=col).value > 70:
        ws.cell(row=22, column=col).value = 0
    else:
        ws.cell(row=22, column=col).value = 0.91 * (ws.cell(row=19, column=col).value - 70)

    if spec_cons < 618:
        ws.cell(row=23, column=col).value = 0
    else:
        ws.cell(row=23, column=col).value = -0.000875 * load_factor * (spec_cons - 618)

    ws.cell(row=24, column=col).value = 92.64 + 1.9475 * load_factor
    ws.cell(row=26, column=col).value = (
        ws.cell(row=24, column=col).value + ws.cell(row=25, column=col).value
    )
    ws.cell(row=27, column=col).value = (
        ws.cell(row=26, column=col).value
        + ws.cell(row=22, column=col).value
        + ws.cell(row=23, column=col).value
    )

    # Коефіцієнт 28
    if load_factor >= 20:
        ws.cell(row=28, column=col).value = 1.17
    else:
        ws.cell(row=28, column=col).value = 1.25 - 0.004 * load_factor

    ws.cell(row=30, column=col).value = ws.cell(row=28, column=col).value * (
        1 + ws.cell(row=29, column=col).value / 100
    )

    # Складна формула рядка 31
    part_a = 3.53 * ws.cell(row=30, column=col).value + 0.6
    part_b = (
        ws.cell(row=27, column=col).value
        - ws.cell(row=30, column=col).value
        * ws.cell(row=21, column=col).value
        / (ws.cell(row=30, column=col).value + 0.18)
    )
    part_c = 0.9805 + 1.3 * ws.cell(row=27, column=col).value / 10000
    ws.cell(row=31, column=col).value = part_a * part_b * part_c / 100

    ws.cell(row=32, column=col).value = 0.04
    ws.cell(row=34, column=col).value = ws.cell(row=32, column=col).value * (
        1 + ws.cell(row=33, column=col).value / 100
    )
    ws.cell(row=35, column=col).value = 0.05
    ws.cell(row=37, column=col).value = ws.cell(row=35, column=col).value * (
        1 + ws.cell(row=36, column=col).value / 100
    )

    ws.cell(row=38, column=col).value = (
        0.00045 * (ws.cell(row=9, column=col).value - ws.cell(row=10, column=col).value) / 1000
    )

    ws.cell(row=39, column=col).value = 100 - (
        ws.cell(row=31, column=col).value
        + ws.cell(row=34, column=col).value
        + ws.cell(row=37, column=col).value
        + ws.cell(row=38, column=col).value
    )

    ws.cell(row=40, column=col).value = (
        ws.cell(row=11, column=col).value / ws.cell(row=39, column=col).value / 7 * 100
    )
    ws.cell(row=41, column=col).value = (
        ws.cell(row=40, column=col).value * 7000 / ws.cell(row=4, column=8).value
    )

    ws.cell(row=42, column=col).value = 2.2354 + 48.9609 / load_factor
    ws.cell(row=44, column=col).value = ws.cell(row=42, column=col).value * (
        1 + ws.cell(row=43, column=col).value / 100
    )
    ws.cell(row=45, column=col).value = (
        ws.cell(row=44, column=col).value * ws.cell(row=11, column=col).value
    )


# =============================================
# ГОЛОВНА ФУНКЦІЯ
# =============================================
def calc_ppks(ws, nmesac: int):
    """Запуск повного розрахунку (аналог VBA CalcPPKs)"""
    # === Обробка трьох ВК ===
    process_ppk_column(ws, 5)   # ВК-1
    process_ppk_column(ws, 6)   # ВК-2
    process_ppk_column(ws, 7)   # ВК-3

    # === Підсумки в стовпці 8 ===
    ws.cell(row=6, column=8).value = (
        ws.cell(row=6, column=5).value
        + ws.cell(row=6, column=6).value
        + ws.cell(row=6, column=7).value
    )
    ws.cell(row=7, column=8).value = (
        ws.cell(row=7, column=5).value
        + ws.cell(row=7, column=6).value
        + ws.cell(row=7, column=7).value
    )
    ws.cell(row=11, column=8).value = (
        ws.cell(row=11, column=5).value
        + ws.cell(row=11, column=6).value
        + ws.cell(row=11, column=7).value
    )
    ws.cell(row=13, column=8).value = (
        ws.cell(row=13, column=5).value
        + ws.cell(row=13, column=6).value
        + ws.cell(row=13, column=7).value
    )
    ws.cell(row=40, column=8).value = (
        ws.cell(row=40, column=5).value
        + ws.cell(row=40, column=6).value
        + ws.cell(row=40, column=7).value
    )
    ws.cell(row=41, column=8).value = (
        ws.cell(row=41, column=5).value
        + ws.cell(row=41, column=6).value
        + ws.cell(row=41, column=7).value
    )
    ws.cell(row=45, column=8).value = (
        ws.cell(row=45, column=5).value
        + ws.cell(row=45, column=6).value
        + ws.cell(row=45, column=7).value
    )

    # === Блок розрахунків після ВК (рядки 48–94) ===
    if ws.cell(row=47, column=8).value == 0:
        for k in range(48, 94):
            ws.cell(row=k, column=8).value = 0
    else:
        ws.cell(row=48, column=8).value = (
            ws.cell(row=11, column=8).value / ws.cell(row=47, column=8).value
        )
        ws.cell(row=50, column=8).value = (
            ws.cell(row=49, column=8).value / ws.cell(row=47, column=8).value
        )
        ws.cell(row=52, column=8).value = (
            ws.cell(row=51, column=8).value / ws.cell(row=47, column=8).value
        )

        ws.cell(row=58, column=8).value = 0.365 * (18 - ws.cell(row=56, column=8).value) / 38
        ws.cell(row=59, column=8).value = 0.0014
        ws.cell(row=60, column=8).value = (
            ws.cell(row=58, column=8).value + ws.cell(row=59, column=8).value
        )
        ws.cell(row=61, column=8).value = (
            ws.cell(row=60, column=8).value / ws.cell(row=48, column=8).value * 100
        )

        ws.cell(row=62, column=8).value = (
            0.00942
            + 0.000247 * (ws.cell(row=54, column=8).value - ws.cell(row=56, column=8).value)
            + 0.0109
            + 0.000343 * (ws.cell(row=55, column=8).value - ws.cell(row=56, column=8).value)
        )

        if ws.cell(row=53, column=8).value == 0:
            ws.cell(row=63, column=8).value = 0
        else:
            if 4 < nmesac < 10:
                ws.cell(row=63, column=8).value = (
                    0.000105 * ws.cell(row=53, column=8).value - 0.0022
                )
            else:
                ws.cell(row=63, column=8).value = (
                    0.00103 * ws.cell(row=53, column=8).value - 0.0212
                )

        ws.cell(row=64, column=8).value = (
            ws.cell(row=62, column=8).value + ws.cell(row=63, column=8).value
        )
        ws.cell(row=65, column=8).value = (
            ws.cell(row=64, column=8).value
            / (
                ws.cell(row=48, column=8).value
                - ws.cell(row=60, column=8).value
                - ws.cell(row=64, column=8).value
            )
            * 100
        )

        # Коефіцієнт 66 залежно від діапазону
        flow = ws.cell(row=52, column=8).value
        if flow <= 500:
            ws.cell(row=66, column=8).value = (
                0.288 + 0.0000745477 * flow - 0.000000421608 * flow**2
            )
        elif flow <= 1240:
            ws.cell(row=66, column=8).value = (
                0.9846
                - 0.0011435 * flow
                + 0.000000848114 * flow**2
                - 0.000000000228469 * flow**3
            )
        elif flow <= 2500:
            ws.cell(row=66, column=8).value = (
                0.7791 - 0.000239328 * flow + 0.0000000403543 * flow**2
            )
        elif flow <= 3750:
            ws.cell(row=66, column=8).value = (
                0.8707
                - 0.000272322 * flow
                + 0.0000000600095 * flow**2
                - 4.91536e-12 * flow**3
            )
        else:
            ws.cell(row=66, column=8).value = 1.1 - 0.000168 * flow

        ws.cell(row=68, column=8).value = ws.cell(row=66, column=8).value * (
            1 + ws.cell(row=67, column=8).value / 100
        )
        ws.cell(row=69, column=8).value = (
            ws.cell(row=68, column=8).value * ws.cell(row=51, column=8).value
        )

        if ws.cell(row=48, column=8).value <= 100:
            ws.cell(row=70, column=8).value = (
                2.1377 - 0.0133 * ws.cell(row=48, column=8).value
            )
        else:
            ws.cell(row=70, column=8).value = (
                3.5094
                - 0.035 * ws.cell(row=48, column=8).value
                + 1.00422 * ws.cell(row=48, column=8).value**2
            )

        ws.cell(row=72, column=8).value = ws.cell(row=70, column=8).value * (
            1 + ws.cell(row=71, column=8).value / 100
        )
        ws.cell(row=73, column=8).value = (
            ws.cell(row=72, column=8).value * ws.cell(row=11, column=8).value
        )

        if ws.cell(row=53, column=8).value == 0:
            ws.cell(row=74, column=8).value = 0
        else:
            ws.cell(row=74, column=8).value = (
                0.2092 + 22.47 / ws.cell(row=53, column=8).value
            )

        ws.cell(row=76, column=8).value = ws.cell(row=74, column=8).value * (
            1 + ws.cell(row=75, column=8).value / 100
        )

        ws.cell(row=77, column=8).value = (
            ws.cell(row=76, column=8).value
            * ws.cell(row=53, column=8).value
            * ws.cell(row=47, column=8).value
        )
        ws.cell(row=78, column=8).value = 35 * ws.cell(row=47, column=8).value
        ws.cell(row=80, column=8).value = ws.cell(row=78, column=8).value * (
            1 + ws.cell(row=79, column=8).value / 100
        )
        ws.cell(row=81, column=8).value = 16.8 * ws.cell(row=47, column=8).value
        ws.cell(row=83, column=8).value = ws.cell(row=81, column=8).value * (
            1 + ws.cell(row=82, column=8).value / 100
        )

        # Складна формула рядка 84
        delta_term = (
            (
                ws.cell(row=9, column=5).value - ws.cell(row=10, column=5).value
                + ws.cell(row=9, column=6).value - ws.cell(row=10, column=6).value
                + ws.cell(row=9, column=7).value - ws.cell(row=10, column=7).value
            )
            / 3
            * 0.0000000045
        )
        ws.cell(row=84, column=8).value = (
            ws.cell(row=45, column=8).value
            + ws.cell(row=69, column=8).value
            + ws.cell(row=73, column=8).value
            + ws.cell(row=77, column=8).value
            + ws.cell(row=80, column=8).value
            + ws.cell(row=83, column=8).value
        ) * (1 + delta_term)

        ws.cell(row=86, column=8).value = (
            ws.cell(row=84, column=8).value / ws.cell(row=49, column=8).value
        )
        ws.cell(row=87, column=8).value = ws.cell(row=86, column=8).value * (
            1 + ws.cell(row=85, column=8).value / 100
        )
        ws.cell(row=88, column=8).value = (
            860 * ws.cell(row=69, column=8).value * 95 / ws.cell(row=49, column=8).value / 1000000
        )

        if ws.cell(row=40, column=8).value == 0:
            ws.cell(row=89, column=8).value = 0
        else:
            ws.cell(row=89, column=8).value = (
                ws.cell(row=39, column=5).value * ws.cell(row=40, column=5).value
                + ws.cell(row=39, column=6).value * ws.cell(row=40, column=6).value
                + ws.cell(row=39, column=7).value * ws.cell(row=40, column=7).value
            ) / ws.cell(row=40, column=8).value

        ws.cell(row=90, column=8).value = ws.cell(row=89, column=8).value * (
            100 - ws.cell(row=61, column=8).value
        ) / 100
        ws.cell(row=91, column=8).value = (
            (100 - ws.cell(row=88, column=8).value + ws.cell(row=65, column=8).value)
            / 7
            / ws.cell(row=90, column=8).value
            * 1000
        )
        ws.cell(row=93, column=8).value = ws.cell(row=91, column=8).value * (
            1 + ws.cell(row=92, column=8).value / 100
        )
        ws.cell(row=94, column=8).value = (
            ws.cell(row=6, column=8).value
            - ws.cell(row=49, column=8).value * ws.cell(row=93, column=8).value / 1000
        )


# =============================================
# ПРИКЛАД ВИКОРИСТАННЯ
# =============================================
if __name__ == "__main__":
    wb = load_workbook("G:\\other\\PTV\\Py_calculation\\data\\exel\\cerkassy_test_ОБНУЛЕНО_20260402_135115.xlsx")          # ← твій файл
    ws = wb["ППК"]                             # ← назва аркуша (зміни при потребі)
    nmesac = 5                                    # ← номер місяця (1-12)

    calc_ppks(ws, nmesac)

    wb.save("G:\\other\\PTV\\Py_calculation\\data\\exel\\cerkassy_test_ppk_calc.xlsx")             # збереження результатів
    print("Розрахунок завершено! Файл збережено як your_file_updated.xlsx")