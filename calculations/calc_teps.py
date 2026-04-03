import openpyxl
from openpyxl import load_workbook


# =============================================
# ГОЛОВНА ФУНКЦІЯ CalcTEPs (повний переклад VBA)
# =============================================
def calc_teps(
    ws_tep,      # аркуш shtTEP
    ws_tur1,     # аркуш shtTur1
    ws_tur2,     # аркуш shtTur2
    ws_kot1,     # аркуш shtKot1
    ws_kot2,     # аркуш shtKot2
    ws_pwk,      # аркуш shtPWK
    nmesac: int  # номер місяця (NMESAC)
):
    """Повний розрахунок TEP — 100% ідентично VBA CalcTEPs()"""

    # ===================== БАЗОВІ РОЗРАХУНКИ =====================
    ws_tep.cell(row=8, column=5).value = ws_tep.cell(row=7, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=10, column=5).value = ws_tep.cell(row=9, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=13, column=5).value = ws_tep.cell(row=12, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=14, column=5).value = ws_tep.cell(row=12, column=5).value / ws_tep.cell(row=9, column=5).value * 100
    ws_tep.cell(row=16, column=5).value = ws_tep.cell(row=15, column=5).value / ws_tep.cell(row=9, column=5).value * 100
    ws_tep.cell(row=18, column=5).value = ws_tep.cell(row=17, column=5).value / ws_tep.cell(row=6, column=5).value

    # ===================== ВИЗНАЧЕННЯ КОЕФІЦІЄНТА 21 =====================
    type_str = ws_tep.cell(row=22, column=5).value
    if type_str == "1С1Н":
        ws_tep.cell(row=21, column=5).value = 0.656 - 0.0001444 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "1С2Н":
        ws_tep.cell(row=21, column=5).value = 0.86 - 0.000148 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "1С3Н":
        ws_tep.cell(row=21, column=5).value = 0.9435 - 0.00013 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "1Ц4Н":
        ws_tep.cell(row=21, column=5).value = 1.074 - 0.000109 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "1Ц5Н":
        ws_tep.cell(row=21, column=5).value = 1.064 - 0.00009 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "6Н1Ц":
        ws_tep.cell(row=21, column=5).value = 1.058 - 0.000077 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "7Н1Ц":
        ws_tep.cell(row=21, column=5).value = 1.039 - 0.000065 * (ws_tep.cell(row=18, column=5).value or 0)
    elif type_str == "8Н1Ц":
        ws_tep.cell(row=21, column=5).value = 1.0683 - 0.0000616666 * (ws_tep.cell(row=18, column=5).value or 0)

    # ===================== ПРОДОВЖЕННЯ РОЗРАХУНКІВ =====================
    ws_tep.cell(row=24, column=5).value = 0.86 * (ws_tep.cell(row=21, column=5).value or 0) * (ws_tep.cell(row=17, column=5).value or 0) * 0.95 / 1000
    ws_tep.cell(row=25, column=5).value = ws_tep.cell(row=24, column=5).value / (ws_tep.cell(row=9, column=5).value or 1) * 100
    ws_tep.cell(row=27, column=5).value = ws_tep.cell(row=26, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=29, column=5).value = ws_tep.cell(row=28, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=31, column=5).value = 0.045 * (ws_tep.cell(row=28, column=5).value or 0)
    ws_tep.cell(row=33, column=5).value = ws_tep.cell(row=32, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=35, column=5).value = ws_tep.cell(row=34, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=36, column=5).value = (ws_tep.cell(row=34, column=5).value or 0) - (ws_tep.cell(row=31, column=5).value or 0)
    ws_tep.cell(row=37, column=5).value = ws_tep.cell(row=36, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=39, column=5).value = ws_tep.cell(row=38, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=40, column=5).value = (ws_tep.cell(row=32, column=5).value or 0) + (ws_tep.cell(row=34, column=5).value or 0) + (ws_tep.cell(row=38, column=5).value or 0)
    ws_tep.cell(row=41, column=5).value = ws_tep.cell(row=40, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=43, column=5).value = ws_tep.cell(row=42, column=5).value / ws_tep.cell(row=6, column=5).value
    ws_tep.cell(row=45, column=5).value = ws_tep.cell(row=44, column=5).value / ws_tep.cell(row=6, column=5).value

    # Рядок 46
    if (ws_tep.cell(row=43, column=5).value or 0) == 0:
        ws_tep.cell(row=46, column=5).value = 0
    else:
        ws_tep.cell(row=46, column=5).value = 43.8004 / (16.9734 + (ws_tep.cell(row=43, column=5).value or 0))

    ws_tep.cell(row=48, column=5).value = (ws_tep.cell(row=46, column=5).value or 0) * (1 + (ws_tep.cell(row=47, column=5).value or 0) / 100)

    # Рядок 49
    if (ws_tep.cell(row=35, column=5).value or 0) == 0:
        ws_tep.cell(row=49, column=5).value = 0
    else:
        x = ws_tep.cell(row=35, column=5).value or 0
        ws_tep.cell(row=49, column=5).value = 1.9803 - 0.0137 * x + 0.0000461048 * x * x

    ws_tep.cell(row=51, column=5).value = (ws_tep.cell(row=49, column=5).value or 0) * (1 + (ws_tep.cell(row=50, column=5).value or 0) / 100)

    # Рядок 52
    if (ws_tep.cell(row=33, column=5).value or 0) == 0:
        ws_tep.cell(row=52, column=5).value = 0
    else:
        x = ws_tep.cell(row=33, column=5).value or 0
        ws_tep.cell(row=52, column=5).value = 1.053 - 0.002573 * x + 0.000003689 * x * x

    ws_tep.cell(row=54, column=5).value = (ws_tep.cell(row=52, column=5).value or 0) * (1 + (ws_tep.cell(row=53, column=5).value or 0) / 100)

    # Рядок 55
    if (ws_tep.cell(row=27, column=5).value or 0) == 0:
        ws_tep.cell(row=55, column=5).value = 0
    else:
        ws_tep.cell(row=55, column=5).value = 0.1041 + 12.0389 / (ws_tep.cell(row=27, column=5).value or 1)

    ws_tep.cell(row=57, column=5).value = (ws_tep.cell(row=55, column=5).value or 0) * (1 + (ws_tep.cell(row=56, column=5).value or 0) / 100)

    # QBOYL (використовується тільки для розрахунку, але в цьому субі не застосовується далі)
    qboyl = (
        (ws_tur1.cell(row=138, column=7).value or 0) - (ws_tur1.cell(row=136, column=7).value or 0) +
        (ws_tur2.cell(row=147, column=8).value or 0) - (ws_tur2.cell(row=145, column=8).value or 0)
    ) / (ws_tep.cell(row=6, column=5).value or 1)

    # Рядок 58 (залежно від діапазону)
    val45 = ws_tep.cell(row=45, column=5).value or 0
    if val45 <= 65:
        ws_tep.cell(row=58, column=5).value = 1.89 - 0.013 * val45
    elif val45 <= 130:
        ws_tep.cell(row=58, column=5).value = 1.9 - 0.00692307 * val45
    elif val45 <= 193:
        ws_tep.cell(row=58, column=5).value = 1.6893 - 0.00357142 * val45
    elif val45 <= 257:
        ws_tep.cell(row=58, column=5).value = 1.6893 - 0.00357142 * val45
    else:
        ws_tep.cell(row=58, column=5).value = 1.6893 - 0.00357142 * val45  # останній діапазон за VBA

    ws_tep.cell(row=60, column=5).value = (ws_tep.cell(row=58, column=5).value or 0) * (1 + (ws_tep.cell(row=59, column=5).value or 0) / 100)

    # Рядок 61
    ws_tep.cell(row=61, column=5).value = (
        (ws_tep.cell(row=54, column=5).value or 0) * (ws_tep.cell(row=32, column=5).value or 0) +
        (ws_tep.cell(row=21, column=5).value or 0) * (ws_tep.cell(row=17, column=5).value or 0) +
        (ws_tep.cell(row=57, column=5).value or 0) * (ws_tep.cell(row=26, column=5).value or 0) +
        (ws_tep.cell(row=60, column=5).value or 0) * (ws_tep.cell(row=44, column=5).value or 0)
    ) / 1000

    # Рядок 62
    if (ws_tep.cell(row=40, column=5).value or 0) == 0:
        ws_tep.cell(row=62, column=5).value = 0
    else:
        ws_tep.cell(row=62, column=5).value = (
            399.9 * (ws_tep.cell(row=32, column=5).value or 0) /
            (ws_tep.cell(row=40, column=5).value or 1) *
            (ws_tep.cell(row=6, column=5).value or 1) / 1000
        )

    ws_tep.cell(row=64, column=5).value = (ws_tep.cell(row=62, column=5).value or 0) * (1 + (ws_tep.cell(row=63, column=5).value or 0) / 100)

    # Рядок 65
    ws_tep.cell(row=65, column=5).value = 60 * (
        (ws_tur1.cell(row=23, column=5).value or 0) + (ws_tur1.cell(row=26, column=5).value or 0) +
        (ws_tur1.cell(row=29, column=5).value or 0) + (ws_tur1.cell(row=32, column=5).value or 0) +
        (ws_tur2.cell(row=23, column=5).value or 0) + (ws_tur2.cell(row=26, column=5).value or 0) +
        (ws_tur2.cell(row=29, column=5).value or 0) + (ws_tur2.cell(row=32, column=5).value or 0) +
        (ws_tur2.cell(row=23, column=6).value or 0) + (ws_tur2.cell(row=26, column=6).value or 0) +
        (ws_tur2.cell(row=29, column=6).value or 0) + (ws_tur2.cell(row=32, column=6).value or 0)
    ) / 1000

    ws_tep.cell(row=67, column=5).value = (ws_tep.cell(row=65, column=5).value or 0) * (1 + (ws_tep.cell(row=66, column=5).value or 0) / 100)

    # Рядок 68–70
    ws_tep.cell(row=68, column=5).value = ws_pwk.cell(row=91, column=10).value
    ws_tep.cell(row=69, column=5).value = (
        (ws_tep.cell(row=61, column=5).value or 0) +
        (ws_tep.cell(row=64, column=5).value or 0) +
        (ws_tep.cell(row=67, column=5).value or 0) +
        (ws_tep.cell(row=68, column=5).value or 0)
    )
    ws_tep.cell(row=70, column=5).value = ws_tep.cell(row=69, column=5).value * 1000 / (ws_tep.cell(row=15, column=5).value or 1)

    # Рядок 71–76
    ws_tep.cell(row=71, column=5).value = (
        (ws_tep.cell(row=48, column=5).value or 0) * (ws_tep.cell(row=42, column=5).value or 0) +
        (ws_tep.cell(row=51, column=5).value or 0) * (ws_tep.cell(row=36, column=5).value or 0)
    ) / 1000

    if (ws_tep.cell(row=40, column=5).value or 0) == 0:
        ws_tep.cell(row=72, column=5).value = 0
    else:
        ws_tep.cell(row=72, column=5).value = (
            399.9 * ((ws_tep.cell(row=36, column=5).value or 0) + (ws_tep.cell(row=42, column=5).value or 0)) /
            1000 / (ws_tep.cell(row=40, column=5).value or 1) * (ws_tep.cell(row=6, column=5).value or 1)
        )

    ws_tep.cell(row=74, column=5).value = (ws_tep.cell(row=72, column=5).value or 0) * (1 + (ws_tep.cell(row=73, column=5).value or 0) / 100)
    ws_tep.cell(row=75, column=5).value = (ws_tep.cell(row=71, column=5).value or 0) + (ws_tep.cell(row=74, column=5).value or 0)

    if (ws_tep.cell(row=7, column=5).value or 0) == 0:
        ws_tep.cell(row=76, column=5).value = 0
    else:
        ws_tep.cell(row=76, column=5).value = ws_tep.cell(row=75, column=5).value * 1000 / ((ws_tep.cell(row=9, column=5).value or 0) - (ws_tep.cell(row=15, column=5).value or 0))

    # ===================== БЛОК ТЕПЛОВИХ ВТРАТ =====================
    ws_tep.cell(row=78, column=5).value = 0.1333 + 0.00333333 * (280 - (ws_kot2.cell(row=27, column=10).value or 0))

    ws_tep.cell(row=77, column=5).value = (
        0.2243 + 0.00389 * ((ws_tep.cell(row=19, column=5).value or 0) - (ws_kot2.cell(row=27, column=10).value or 0)) +
        0.1733 + 0.00233333 * ((ws_tep.cell(row=20, column=5).value or 0) - (ws_kot2.cell(row=27, column=10).value or 0))
    )

    ws_tep.cell(row=79, column=5).value = (
        0.241 - 0.0172 * (ws_kot2.cell(row=27, column=10).value or 0) +
        0.2198 - 0.01576 * (ws_kot2.cell(row=27, column=10).value or 0) +
        0.1836 - 0.0133 * (ws_kot2.cell(row=27, column=10).value or 0)
    )
    if (ws_tep.cell(row=79, column=5).value or 0) < 0:
        ws_tep.cell(row=79, column=5).value = 0

    ws_tep.cell(row=80, column=5).value = 1.924 if 4 < nmesac < 10 else 2.918

    ws_tep.cell(row=81, column=5).value = (
        (ws_tep.cell(row=77, column=5).value or 0) +
        (ws_tep.cell(row=78, column=5).value or 0) +
        (ws_tep.cell(row=79, column=5).value or 0) +
        (ws_tep.cell(row=80, column=5).value or 0)
    )
    ws_tep.cell(row=82, column=5).value = ws_tep.cell(row=81, column=5).value / (ws_tep.cell(row=10, column=5).value or 1) * 100
    ws_tep.cell(row=84, column=5).value = (ws_tep.cell(row=82, column=5).value or 0) * (1 + (ws_tep.cell(row=83, column=5).value or 0) / 100)

    # ===================== РЯДКИ 96–110 =====================
    ws_tep.cell(row=96, column=5).value = (
        ws_tur1.cell(row=7, column=5).value * ws_tur1.cell(row=130, column=5).value +
        ws_tur1.cell(row=7, column=6).value * ws_tur1.cell(row=130, column=6).value +
        ws_tur2.cell(row=7, column=5).value * ws_tur2.cell(row=139, column=5).value +
        ws_tur2.cell(row=7, column=6).value * ws_tur2.cell(row=139, column=6).value +
        ws_tur2.cell(row=7, column=7).value * ws_tur2.cell(row=139, column=7).value
    ) / (ws_tep.cell(row=87, column=5).value or 1)

    ws_tep.cell(row=97, column=5).value = (
        ws_kot1.cell(row=110, column=5).value * ws_kot1.cell(row=106, column=5).value +
        ws_kot1.cell(row=110, column=6).value * ws_kot1.cell(row=106, column=6).value +
        ws_kot1.cell(row=110, column=7).value * ws_kot1.cell(row=106, column=7).value +
        ws_kot1.cell(row=110, column=8).value * ws_kot1.cell(row=106, column=8).value +
        ws_kot2.cell(row=116, column=5).value * ws_kot2.cell(row=112, column=5).value +
        ws_kot2.cell(row=116, column=6).value * ws_kot2.cell(row=112, column=6).value +
        ws_kot2.cell(row=116, column=7).value * ws_kot2.cell(row=112, column=7).value +
        ws_kot2.cell(row=116, column=8).value * ws_kot2.cell(row=112, column=8).value +
        ws_kot2.cell(row=116, column=9).value * ws_kot2.cell(row=112, column=9).value
    ) / ((ws_kot1.cell(row=110, column=9).value or 0) + (ws_kot2.cell(row=116, column=10).value or 0))

    ws_tep.cell(row=99, column=5).value = ws_tur1.cell(row=157, column=7).value + ws_tur2.cell(row=167, column=8).value

    # Рядок 100 (залежно від типу)
    type100 = ws_tep.cell(row=101, column=5).value
    x29 = ws_tep.cell(row=29, column=5).value or 0
    if type100 == "1М":
        ws_tep.cell(row=100, column=5).value = 15.08321 - 0.112044 * x29 + 0.000336 * x29 * x29
    elif type100 == "1В":
        ws_tep.cell(row=100, column=5).value = 13.2214 - 0.0443 * x29 + 0.0000611047 * x29 * x29
    elif type100 == "1В1М":
        ws_tep.cell(row=100, column=5).value = 13.9939 - 0.0333 * x29 + 0.0000326697 * x29 * x29
    elif type100 == "2В":
        ws_tep.cell(row=100, column=5).value = 10.0894 - 0.00821075 * x29
    elif type100 == "2В1М":
        ws_tep.cell(row=100, column=5).value = 12.8599 - 0.0164 * x29 + 0.00000880624 * x29 * x29
    elif type100 == "3В":
        ws_tep.cell(row=100, column=5).value = 18.0464 - 0.0272 * x29 + 0.0000148107 * x29 * x29
    elif type100 == "3В1М":
        ws_tep.cell(row=100, column=5).value = 15.3563 - 0.017 * x29 + 0.00000730989 * x29 * x29
    elif type100 == "4В":
        ws_tep.cell(row=100, column=5).value = 18.6276 - 0.0215 * x29 + 0.00000882842 * x29 * x29
    elif type100 == "4В1М":
        ws_tep.cell(row=100, column=5).value = 18.3105 - 0.0183 * x29 + 0.00000660677 * x29 * x29
    elif type100 == "5В":
        ws_tep.cell(row=100, column=5).value = 21.4296 - 0.0214 * x29 + 0.00000224511 * x29 * x29
    elif type100 == "5В1М":
        ws_tep.cell(row=100, column=5).value = 21.5352 - 0.0195 * x29 + 0.00000597756 * x29 * x29

    ws_tep.cell(row=103, column=5).value = (ws_tep.cell(row=100, column=5).value or 0) * (ws_tep.cell(row=28, column=5).value or 0) / 1000
    ws_tep.cell(row=104, column=5).value = (16.7 + 42.9 + 504) * (ws_tep.cell(row=6, column=5).value or 0) / 1000
    ws_tep.cell(row=106, column=5).value = (ws_tep.cell(row=104, column=5).value or 0) + (ws_tep.cell(row=105, column=5).value or 0) * (ws_tep.cell(row=6, column=5).value or 0) / 1000

    ws_tep.cell(row=112, column=5).value = ws_kot1.cell(row=130, column=9).value + ws_kot2.cell(row=136, column=10).value

    x35 = ws_tep.cell(row=35, column=5).value or 0
    ws_tep.cell(row=113, column=5).value = 1.9803 - 0.0137 * x35 + 0.0000461048 * x35 * x35
    ws_tep.cell(row=115, column=5).value = (ws_tep.cell(row=113, column=5).value or 0) * (1 + (ws_tep.cell(row=114, column=5).value or 0) / 100)

    if (ws_tep.cell(row=40, column=5).value or 0) == 0:
        ws_tep.cell(row=116, column=5).value = 0
    else:
        ws_tep.cell(row=116, column=5).value = (
            (ws_tep.cell(row=115, column=5).value or 0) * (ws_tep.cell(row=31, column=5).value or 0) / 1000 +
            399.9 * (ws_tep.cell(row=31, column=5).value or 0) / (ws_tep.cell(row=40, column=5).value or 1) *
            (ws_tep.cell(row=6, column=5).value or 0) / 1000
        )

    # Рядок 117
    val90 = ws_tep.cell(row=90, column=5).value or 0
    if val90 > 1:
        ws_tep.cell(row=117, column=5).value = (
            0.39 * (ws_tep.cell(row=89, column=5).value or 0) +
            0.78 * val90 +
            60 * (ws_tep.cell(row=91, column=5).value or 0) / 60 / 16 +
            67.5 * (ws_tep.cell(row=94, column=5).value or 0)
        ) / 1000
    else:
        ws_tep.cell(row=117, column=5).value = (
            0.39 * (ws_tep.cell(row=89, column=5).value or 0) +
            0.78 * val90 +
            60 * (ws_tep.cell(row=91, column=5).value or 0) / 60 / 16 +
            45 * (ws_tep.cell(row=94, column=5).value or 0)
        ) / 1000

    ws_tep.cell(row=119, column=5).value = (ws_tep.cell(row=117, column=5).value or 0) * (1 + (ws_tep.cell(row=118, column=5).value or 0) / 100)

    # Рядок 120
    if val90 < 1:
        ws_tep.cell(row=120, column=5).value = 345 * (ws_tep.cell(row=6, column=5).value or 0) / 1000
    else:
        su1 = 402 * (ws_tep.cell(row=6, column=5).value or 0) / 1000 if (ws_kot1.cell(row=111, column=9).value or 0) > 1 else 345 * (ws_tep.cell(row=6, column=5).value or 0) / 1000
        su2 = 902.9 * (ws_tep.cell(row=6, column=5).value or 0) / 1000 if (ws_kot2.cell(row=117, column=10).value or 0) > 1 else 0
        ws_tep.cell(row=120, column=5).value = su1 + su2

    # Рядки 123–126
    ws_tep.cell(row=123, column=5).value = (
        (
            (ws_tep.cell(row=112, column=5).value or 0) +
            (ws_tep.cell(row=116, column=5).value or 0) +
            (ws_tep.cell(row=119, column=5).value or 0) +
            (ws_tep.cell(row=120, column=5).value or 0) +
            (ws_tep.cell(row=122, column=5).value or 0) * (ws_tep.cell(row=6, column=5).value or 0) / 1000
        ) * (1 + (ws_tep.cell(row=96, column=5).value or 0) / 100)
    )
    ws_tep.cell(row=124, column=5).value = ws_tep.cell(row=123, column=5).value / (ws_tep.cell(row=87, column=5).value or 1) * 100

    ws_tep.cell(row=107, column=5).value = (
        (ws_tep.cell(row=95, column=5).value or 0) +
        (ws_tep.cell(row=99, column=5).value or 0) +
        (ws_tep.cell(row=103, column=5).value or 0) +
        (ws_tep.cell(row=106, column=5).value or 0)
    ) * (1 + (ws_tep.cell(row=96, column=5).value or 0) / 100)
    ws_tep.cell(row=108, column=5).value = ws_tep.cell(row=107, column=5).value / (ws_tep.cell(row=87, column=5).value or 1) * 100

    ws_tep.cell(row=109, column=5).value = ws_tur1.cell(row=150, column=7).value + ws_tur2.cell(row=160, column=8).value
    ws_tep.cell(row=110, column=5).value = ws_tep.cell(row=109, column=5).value / ((ws_tur1.cell(row=136, column=7).value or 0) + (ws_tur2.cell(row=145, column=8).value or 0)) * 100

    ws_tep.cell(row=125, column=5).value = (ws_tep.cell(row=107, column=5).value or 0) + (ws_tep.cell(row=123, column=5).value or 0) + (ws_tep.cell(row=69, column=5).value or 0) + (ws_tep.cell(row=75, column=5).value or 0)
    ws_tep.cell(row=126, column=5).value = ws_tep.cell(row=125, column=5).value / (ws_tep.cell(row=87, column=5).value or 1) * 100

    # ===================== ФІНАЛЬНИЙ БЛОК (128–170) =====================
    ws_tep.cell(row=128, column=5).value = ws_kot1.cell(row=138, column=9).value + ws_kot2.cell(row=146, column=10).value

    kot_temp = ws_kot2.cell(row=27, column=10).value or 0
    if kot_temp <= 8:
        ws_tep.cell(row=129, column=5).value = (ws_tep.cell(row=6, column=5).value or 0) * (
            1.1514 - 0.07893 * (ws_kot1.cell(row=27, column=9).value or 0) +
            0.7228 - 0.04286 * kot_temp
        )
    else:
        ws_tep.cell(row=129, column=5).value = (ws_tep.cell(row=6, column=5).value or 0) * (
            0.1 - 0.005 * (ws_kot1.cell(row=27, column=9).value or 0) +
            0.22 - 0.01 * kot_temp
        )
    if (ws_tep.cell(row=129, column=5).value or 0) < 0:
        ws_tep.cell(row=129, column=5).value = 0

    ws_tep.cell(row=131, column=5).value = (ws_tep.cell(row=129, column=5).value or 0) * (1 + (ws_tep.cell(row=130, column=5).value or 0) / 100)

    pwk67 = ws_pwk.cell(row=67, column=10).value or 0
    if pwk67 == 0:
        if kot_temp <= 0:
            ws_tep.cell(row=132, column=5).value = (ws_tep.cell(row=94, column=5).value or 0) * 2.45
        else:
            ws_tep.cell(row=132, column=5).value = (ws_tep.cell(row=94, column=5).value or 0) * 2.15
    else:
        x = kot_temp + 20
        ws_tep.cell(row=132, column=5).value = pwk67 * (0.1325 - 0.0010438 * x - 0.00000902164 * x * x)

    ws_tep.cell(row=134, column=5).value = (ws_tep.cell(row=132, column=5).value or 0) * (1 + (ws_tep.cell(row=133, column=5).value or 0) / 100)

    ws_tep.cell(row=135, column=5).value = (ws_tep.cell(row=28, column=5).value or 0) * (1.4 if 4 < nmesac < 10 else 2.34) / 10000
    ws_tep.cell(row=137, column=5).value = (ws_tep.cell(row=135, column=5).value or 0) * (1 + (ws_tep.cell(row=136, column=5).value or 0) / 100)

    ws_tep.cell(row=138, column=5).value = (
        (ws_tep.cell(row=91, column=5).value or 0) *
        (12.0128 - 0.6354 * (kot_temp + 30) + 0.0104 * (kot_temp + 30) * (kot_temp + 30))
    ) / 1000
    ws_tep.cell(row=140, column=5).value = (ws_tep.cell(row=138, column=5).value or 0) * (1 + (ws_tep.cell(row=139, column=5).value or 0) / 100)

    ws_tep.cell(row=141, column=5).value = (
        (ws_tep.cell(row=92, column=5).value or 0) *
        (1.2388 - 0.0114 * (kot_temp + 30) + 0.000150855 * (kot_temp + 30) * (kot_temp + 30))
    )
    ws_tep.cell(row=143, column=5).value = (ws_tep.cell(row=141, column=5).value or 0) * (1 + (ws_tep.cell(row=142, column=5).value or 0) / 100)

    ws_tep.cell(row=144, column=5).value = (ws_tep.cell(row=140, column=5).value or 0) + (ws_tep.cell(row=143, column=5).value or 0)
    ws_tep.cell(row=145, column=5).value = (
        (ws_tep.cell(row=128, column=5).value or 0) +
        (ws_tep.cell(row=131, column=5).value or 0) +
        (ws_tep.cell(row=134, column=5).value or 0) +
        (ws_tep.cell(row=137, column=5).value or 0) +
        (ws_tep.cell(row=144, column=5).value or 0)
    )
    ws_tep.cell(row=146, column=5).value = ws_tep.cell(row=145, column=5).value / ((ws_kot1.cell(row=16, column=9).value or 0) + (ws_kot2.cell(row=16, column=10).value or 0)) * 100

    ws_tep.cell(row=148, column=5).value = ((ws_tep.cell(row=125, column=5).value or 0) - (ws_tep.cell(row=147, column=5).value or 0)) / (ws_tep.cell(row=125, column=5).value or 1)

    # ===================== СУМА ДЛЯ 149 =====================
    sum149 = (
        # shtTur1
        (ws_tur1.cell(row=36, column=5).value or 0) * (ws_tur1.cell(row=11, column=5).value or 0) +
        (ws_tur1.cell(row=36, column=6).value or 0) * (ws_tur1.cell(row=11, column=6).value or 0) +
        (ws_tur1.cell(row=37, column=5).value or 0) * (ws_tur1.cell(row=14, column=5).value or 0) +
        (ws_tur1.cell(row=38, column=5).value or 0) * (ws_tur1.cell(row=23, column=5).value or 0) +
        (ws_tur1.cell(row=39, column=5).value or 0) * (ws_tur1.cell(row=26, column=5).value or 0) +
        (ws_tur1.cell(row=42, column=5).value or 0) * (ws_tur1.cell(row=11, column=5).value or 0) +
        (ws_tur1.cell(row=42, column=6).value or 0) * (ws_tur1.cell(row=11, column=6).value or 0) +
        (ws_tur1.cell(row=43, column=5).value or 0) * (ws_tur1.cell(row=17, column=5).value or 0) +
        (ws_tur1.cell(row=43, column=6).value or 0) * (ws_tur1.cell(row=17, column=6).value or 0) +
        (ws_tur1.cell(row=44, column=5).value or 0) * (ws_tur1.cell(row=23, column=5).value or 0) +
        (ws_tur1.cell(row=45, column=5).value or 0) * (ws_tur1.cell(row=29, column=5).value or 0) +
        (ws_tur1.cell(row=50, column=5).value or 0) * (
            (ws_tur1.cell(row=23, column=5).value or 0) + (ws_tur1.cell(row=26, column=5).value or 0) +
            (ws_tur1.cell(row=29, column=5).value or 0) + (ws_tur1.cell(row=32, column=5).value or 0)
        )
    )

    # shtTur2
    sum149 += (
        (ws_tur2.cell(row=36, column=5).value or 0) * (ws_tur2.cell(row=11, column=5).value or 0) +
        (ws_tur2.cell(row=36, column=6).value or 0) * (ws_tur2.cell(row=11, column=6).value or 0) +
        (ws_tur2.cell(row=36, column=7).value or 0) * (ws_tur2.cell(row=11, column=7).value or 0) +
        (ws_tur2.cell(row=37, column=5).value or 0) * (ws_tur2.cell(row=14, column=5).value or 0) +
        (ws_tur2.cell(row=37, column=6).value or 0) * (ws_tur2.cell(row=14, column=6).value or 0) +
        (ws_tur2.cell(row=37, column=7).value or 0) * (ws_tur2.cell(row=14, column=7).value or 0) +
        (ws_tur2.cell(row=38, column=5).value or 0) * (ws_tur2.cell(row=23, column=5).value or 0) +
        (ws_tur2.cell(row=38, column=6).value or 0) * (ws_tur2.cell(row=23, column=6).value or 0) +
        (ws_tur2.cell(row=39, column=5).value or 0) * (ws_tur2.cell(row=26, column=5).value or 0) +
        (ws_tur2.cell(row=39, column=6).value or 0) * (ws_tur2.cell(row=26, column=6).value or 0) +
        (ws_tur2.cell(row=44, column=5).value or 0) * (ws_tur2.cell(row=11, column=5).value or 0) +
        (ws_tur2.cell(row=44, column=6).value or 0) * (ws_tur2.cell(row=11, column=6).value or 0) +
        (ws_tur2.cell(row=44, column=7).value or 0) * (ws_tur2.cell(row=11, column=7).value or 0) +
        (ws_tur2.cell(row=45, column=5).value or 0) * (ws_tur2.cell(row=17, column=5).value or 0) +
        (ws_tur2.cell(row=45, column=6).value or 0) * (ws_tur2.cell(row=17, column=6).value or 0) +
        (ws_tur2.cell(row=45, column=7).value or 0) * (ws_tur2.cell(row=17, column=7).value or 0) +
        (ws_tur2.cell(row=46, column=5).value or 0) * (ws_tur2.cell(row=23, column=5).value or 0) +
        (ws_tur2.cell(row=46, column=6).value or 0) * (ws_tur2.cell(row=23, column=6).value or 0) +
        (ws_tur2.cell(row=47, column=5).value or 0) * (ws_tur2.cell(row=29, column=5).value or 0) +
        (ws_tur2.cell(row=47, column=6).value or 0) * (ws_tur2.cell(row=29, column=6).value or 0) +
        (ws_tur2.cell(row=52, column=5).value or 0) * (
            (ws_tur2.cell(row=23, column=5).value or 0) + (ws_tur2.cell(row=26, column=5).value or 0) +
            (ws_tur2.cell(row=29, column=5).value or 0) + (ws_tur2.cell(row=32, column=5).value or 0)
        ) +
        (ws_tur2.cell(row=52, column=6).value or 0) * (
            (ws_tur2.cell(row=23, column=6).value or 0) + (ws_tur2.cell(row=26, column=6).value or 0) +
            (ws_tur2.cell(row=29, column=6).value or 0) + (ws_tur2.cell(row=32, column=6).value or 0)
        )
    )

    ws_tep.cell(row=149, column=5).value = sum149

    ws_tep.cell(row=150, column=5).value = (
        (ws_tep.cell(row=149, column=5).value or 0) -
        (ws_tep.cell(row=109, column=5).value or 0) -
        (ws_tep.cell(row=145, column=5).value or 0) -
        (ws_tep.cell(row=81, column=5).value or 0) * (ws_tep.cell(row=6, column=5).value or 0)
    )

    # ===================== СУМА ДЛЯ 151 (дуже довга частина) =====================
    sum151 = (
        # shtTur1 частина
        (ws_tur1.cell(row=36, column=5).value or 0) * (ws_tur1.cell(row=11, column=5).value or 0) * (1 - (ws_tur1.cell(row=143, column=5).value or 0)) +
        (ws_tur1.cell(row=36, column=6).value or 0) * (ws_tur1.cell(row=11, column=6).value or 0) * (1 - (ws_tur1.cell(row=143, column=6).value or 0)) +
        (ws_tur1.cell(row=37, column=5).value or 0) * (ws_tur1.cell(row=14, column=5).value or 0) * (1 - (ws_tur1.cell(row=143, column=5).value or 0)) +
        (ws_tur1.cell(row=38, column=5).value or 0) * (ws_tur1.cell(row=23, column=5).value or 0) * (1 - (ws_tur1.cell(row=143, column=5).value or 0)) +
        (ws_tur1.cell(row=39, column=5).value or 0) * (ws_tur1.cell(row=26, column=5).value or 0) * (1 - (ws_tur1.cell(row=143, column=5).value or 0)) +
        (ws_tur1.cell(row=42, column=5).value or 0) * (ws_tur1.cell(row=11, column=5).value or 0) * (1 - (ws_tur1.cell(row=144, column=5).value or 0)) +
        (ws_tur1.cell(row=42, column=6).value or 0) * (ws_tur1.cell(row=11, column=6).value or 0) * (1 - (ws_tur1.cell(row=144, column=6).value or 0)) +
        (ws_tur1.cell(row=43, column=5).value or 0) * (ws_tur1.cell(row=17, column=5).value or 0) * (1 - (ws_tur1.cell(row=144, column=5).value or 0)) +
        (ws_tur1.cell(row=43, column=6).value or 0) * (ws_tur1.cell(row=17, column=6).value or 0) * (1 - (ws_tur1.cell(row=144, column=6).value or 0)) +
        (ws_tur1.cell(row=50, column=5).value or 0) * (
            (ws_tur1.cell(row=23, column=5).value or 0) + (ws_tur1.cell(row=26, column=5).value or 0) +
            (ws_tur1.cell(row=29, column=5).value or 0) + (ws_tur1.cell(row=32, column=5).value or 0)
        ) * (1 - (ws_tur1.cell(row=145, column=5).value or 0))
    )

    # shtTur2 частина (змішані посилання — точно як у VBA)
    sum151 += (
        (ws_tur2.cell(row=36, column=5).value or 0) * (ws_tur1.cell(row=11, column=5).value or 0) * (1 - (ws_tur2.cell(row=152, column=5).value or 0)) +
        (ws_tur2.cell(row=36, column=6).value or 0) * (ws_tur2.cell(row=11, column=6).value or 0) * (1 - (ws_tur2.cell(row=152, column=6).value or 0)) +
        (ws_tur2.cell(row=36, column=7).value or 0) * (ws_tur2.cell(row=11, column=7).value or 0) * (1 - (ws_tur2.cell(row=152, column=7).value or 0)) +
        (ws_tur2.cell(row=38, column=5).value or 0) * (ws_tur2.cell(row=23, column=5).value or 0) * (1 - (ws_tur2.cell(row=152, column=5).value or 0)) +
        (ws_tur2.cell(row=38, column=6).value or 0) * (ws_tur2.cell(row=23, column=6).value or 0) * (1 - (ws_tur2.cell(row=152, column=6).value or 0)) +
        (ws_tur2.cell(row=37, column=5).value or 0) * (ws_tur2.cell(row=14, column=5).value or 0) * (1 - (ws_tur2.cell(row=152, column=5).value or 0)) +
        (ws_tur2.cell(row=37, column=6).value or 0) * (ws_tur2.cell(row=14, column=6).value or 0) * (1 - (ws_tur2.cell(row=152, column=6).value or 0)) +
        (ws_tur2.cell(row=37, column=7).value or 0) * (ws_tur2.cell(row=14, column=7).value or 0) * (1 - (ws_tur2.cell(row=152, column=7).value or 0)) +
        (ws_tur2.cell(row=39, column=5).value or 0) * (ws_tur2.cell(row=26, column=5).value or 0) * (1 - (ws_tur2.cell(row=152, column=5).value or 0)) +
        (ws_tur2.cell(row=39, column=6).value or 0) * (ws_tur2.cell(row=26, column=6).value or 0) * (1 - (ws_tur2.cell(row=152, column=6).value or 0)) +
        (ws_tur2.cell(row=44, column=5).value or 0) * (ws_tur2.cell(row=11, column=5).value or 0) * (1 - (ws_tur2.cell(row=153, column=5).value or 0)) +
        (ws_tur2.cell(row=44, column=6).value or 0) * (ws_tur2.cell(row=11, column=6).value or 0) * (1 - (ws_tur2.cell(row=153, column=6).value or 0)) +
        (ws_tur2.cell(row=44, column=7).value or 0) * (ws_tur2.cell(row=11, column=7).value or 0) * (1 - (ws_tur2.cell(row=153, column=7).value or 0)) +
        (ws_tur2.cell(row=46, column=5).value or 0) * (ws_tur2.cell(row=23, column=5).value or 0) * (1 - (ws_tur2.cell(row=153, column=5).value or 0)) +
        (ws_tur2.cell(row=46, column=6).value or 0) * (ws_tur2.cell(row=23, column=6).value or 0) * (1 - (ws_tur2.cell(row=153, column=6).value or 0)) +
        (ws_tur2.cell(row=45, column=5).value or 0) * (ws_tur2.cell(row=17, column=5).value or 0) * (1 - (ws_tur2.cell(row=153, column=5).value or 0)) +
        (ws_tur2.cell(row=45, column=6).value or 0) * (ws_tur2.cell(row=17, column=6).value or 0) * (1 - (ws_tur2.cell(row=153, column=6).value or 0)) +
        (ws_tur2.cell(row=45, column=7).value or 0) * (ws_tur2.cell(row=17, column=7).value or 0) * (1 - (ws_tur2.cell(row=153, column=7).value or 0)) +
        (ws_tur2.cell(row=47, column=5).value or 0) * (ws_tur2.cell(row=29, column=5).value or 0) * (1 - (ws_tur2.cell(row=153, column=5).value or 0)) +
        (ws_tur2.cell(row=47, column=6).value or 0) * (ws_tur2.cell(row=29, column=6).value or 0) * (1 - (ws_tur2.cell(row=153, column=6).value or 0)) +
        (ws_tur2.cell(row=52, column=5).value or 0) * (
            (ws_tur2.cell(row=23, column=5).value or 0) + (ws_tur2.cell(row=26, column=5).value or 0) +
            (ws_tur2.cell(row=29, column=5).value or 0) + (ws_tur2.cell(row=32, column=5).value or 0)
        ) * (1 - (ws_tur2.cell(row=154, column=5).value or 0)) +
        (ws_tur2.cell(row=52, column=6).value or 0) * (
            (ws_tur2.cell(row=23, column=6).value or 0) + (ws_tur2.cell(row=26, column=6).value or 0) +
            (ws_tur2.cell(row=29, column=6).value or 0) + (ws_tur2.cell(row=32, column=6).value or 0)
        ) * (1 - (ws_tur2.cell(row=154, column=6).value or 0))
    )

    val149 = ws_tep.cell(row=149, column=5).value or 1
    ws_tep.cell(row=151, column=5).value = sum151 * (ws_tep.cell(row=150, column=5).value or 0) / val149

    ws_tep.cell(row=152, column=5).value = ws_tur1.cell(row=136, column=7).value + ws_tur2.cell(row=145, column=8).value

    # Рядок 153
    kot_sum = (ws_kot1.cell(row=111, column=9).value or 0) + (ws_kot2.cell(row=117, column=10).value or 0)
    if kot_sum < 0.1:
        ws_tep.cell(row=153, column=5).value = 1
    else:
        ws_tep.cell(row=153, column=5).value = (
            1 * (ws_kot1.cell(row=111, column=9).value or 0) +
            (ws_kot2.cell(row=48, column=5).value or 0) * (ws_kot2.cell(row=117, column=5).value or 0) +
            (ws_kot2.cell(row=48, column=6).value or 0) * (ws_kot2.cell(row=117, column=6).value or 0) +
            (ws_kot2.cell(row=48, column=7).value or 0) * (ws_kot2.cell(row=117, column=7).value or 0) +
            (ws_kot2.cell(row=48, column=8).value or 0) * (ws_kot2.cell(row=117, column=8).value or 0) +
            (ws_kot2.cell(row=48, column=9).value or 0) * (ws_kot2.cell(row=117, column=9).value or 0)
        ) / kot_sum

    ws_tep.cell(row=154, column=5).value = (
        (ws_kot1.cell(row=113, column=9).value or 0) * (ws_kot1.cell(row=110, column=9).value or 0) +
        (ws_kot2.cell(row=119, column=10).value or 0) * (ws_kot2.cell(row=116, column=10).value or 0)
    ) / ((ws_kot1.cell(row=110, column=9).value or 0) + (ws_kot2.cell(row=116, column=10).value or 0))

    ws_tep.cell(row=155, column=5).value = (
        (ws_tur1.cell(row=135, column=7).value or 0) * (ws_tur1.cell(row=7, column=7).value or 0) +
        (ws_tur2.cell(row=144, column=8).value or 0) * (ws_tur2.cell(row=7, column=8).value or 0)
    ) / ((ws_tur1.cell(row=7, column=7).value or 0) + (ws_tur2.cell(row=7, column=8).value or 0))

    ws_tep.cell(row=156, column=5).value = (
        (ws_tep.cell(row=152, column=5).value or 0) + (ws_tep.cell(row=109, column=5).value or 0) + 0.5 * (ws_tep.cell(row=151, column=5).value or 0)
    ) / (
        (ws_tep.cell(row=152, column=5).value or 0) + (ws_tep.cell(row=109, column=5).value or 0) + 0.5 * (ws_tep.cell(row=151, column=5).value or 0) +
        (ws_tep.cell(row=9, column=5).value or 0) *
        (100 - (ws_tep.cell(row=25, column=5).value or 0) - (ws_tep.cell(row=14, column=5).value or 0) +
         (ws_tep.cell(row=84, column=5).value or 0) - (ws_tep.cell(row=11, column=5).value or 0) * 100 / (ws_tep.cell(row=9, column=5).value or 1)) / 100
    )

    # Рядок 157
    ws_tep.cell(row=157, column=5).value = (
        (ws_tep.cell(row=154, column=5).value or 0) *
        (100 - (ws_tep.cell(row=146, column=5).value or 0)) /
        (ws_tep.cell(row=153, column=5).value or 1) /
        (100 - (ws_tep.cell(row=108, column=5).value or 0) * (ws_tep.cell(row=148, column=5).value or 0)) *
        (100 - ((ws_tep.cell(row=108, column=5).value or 0) + (ws_tep.cell(row=156, column=5).value or 0) * (ws_tep.cell(row=124, column=5).value or 0)) * (ws_tep.cell(row=148, column=5).value or 0)) / 100
    )

    # Рядок 158
    kot_loss = (
        64.84 * (
            (ws_kot1.cell(row=17, column=5).value or 0) + (ws_kot1.cell(row=17, column=6).value or 0) +
            (ws_kot1.cell(row=17, column=7).value or 0) + (ws_kot1.cell(row=17, column=8).value or 0)
        ) +
        129.4 * (
            (ws_kot2.cell(row=17, column=5).value or 0) + (ws_kot2.cell(row=17, column=6).value or 0) +
            (ws_kot2.cell(row=17, column=7).value or 0) + (ws_kot2.cell(row=17, column=8).value or 0) +
            (ws_kot2.cell(row=17, column=9).value or 0)
        )
    )
    denom158 = (ws_kot1.cell(row=16, column=9).value or 0) + (ws_kot2.cell(row=16, column=10).value or 0) - (ws_tep.cell(row=145, column=5).value or 0)
    ws_tep.cell(row=158, column=5).value = (1 - 0.015 * kot_loss / (denom158 or 1)) * 100

    ws_tep.cell(row=159, column=5).value = (ws_tep.cell(row=155, column=5).value or 0) * (100 + (ws_tep.cell(row=110, column=5).value or 0)) / (100 - (ws_tep.cell(row=108, column=5).value or 0) * (ws_tep.cell(row=148, column=5).value or 0))
    ws_tep.cell(row=160, column=5).value = ((ws_tep.cell(row=152, column=5).value or 0) + (ws_tep.cell(row=109, column=5).value or 0) + 0.5 * (ws_tep.cell(row=151, column=5).value or 0)) / ((ws_tep.cell(row=152, column=5).value or 0) + (ws_tep.cell(row=109, column=5).value or 0))

    ws_tep.cell(row=161, column=5).value = (
        ((ws_kot1.cell(row=16, column=9).value or 0) + (ws_kot2.cell(row=16, column=10).value or 0) - (ws_tep.cell(row=145, column=5).value or 0)) *
        (ws_tep.cell(row=158, column=5).value or 0) / 100 +
        0.5 * (ws_tep.cell(row=151, column=5).value or 0)
    ) / (
        ((ws_kot1.cell(row=16, column=9).value or 0) + (ws_kot2.cell(row=16, column=10).value or 0) - (ws_tep.cell(row=145, column=5).value or 0)) *
        (ws_tep.cell(row=158, column=5).value or 0) / 100
    ) * 100

    # ===================== КОЕФІЦІЄНТИ kst1 / kst2 =====================
    kst1 = 0.0
    if (ws_kot1.cell(row=16, column=9).value or 0) >= 0.1:
        qkotn1 = (ws_kot1.cell(row=16, column=9).value or 0) / (
            (ws_kot1.cell(row=17, column=5).value or 0) + (ws_kot1.cell(row=17, column=6).value or 0) +
            (ws_kot1.cell(row=17, column=7).value or 0) + (ws_kot1.cell(row=17, column=8).value or 0)
        ) / 70.7

        if qkotn1 >= 0.75:
            x1 = 15.05 - 29.4388 * qkotn1 + 14.3971 * qkotn1 * qkotn1
            x2 = 27.43 - 54.3532 * qkotn1 + 26.9242 * qkotn1 * qkotn1
            kst1u = x1 + (x2 - x1) / 0.25 * (qkotn1 - 0.75)
        else:
            x1 = 11.22 - 21.9621 * qkotn1 + 10.7443 * qkotn1 * qkotn1
            x2 = 15.05 - 29.4388 * qkotn1 + 14.3971 * qkotn1 * qkotn1
            kst1u = x1 + (x2 - x1) / 0.25 * (qkotn1 - 0.5)

        if qkotn1 >= 0.75:
            x1 = 1.78 - 3.5413 * qkotn1 + 1.7596 * qkotn1 * qkotn1
            x2 = 2.24 - 4.4073 * qkotn1 + 2.1738 * qkotn1 * qkotn1
            kst1g = x1 + (x2 - x1) / 0.25 * (qkotn1 - 0.75)
        else:
            x1 = 1.22 - 2.4659 * qkotn1 + 1.2484 * qkotn1 * qkotn1
            x2 = 1.78 - 3.5413 * qkotn1 + 1.7596 * qkotn1 * qkotn1
            kst1g = x1 + (x2 - x1) / 0.25 * (qkotn1 - 0.5)

        kst1 = kst1u * (ws_kot1.cell(row=21, column=9).value or 0) + kst1g * (ws_kot1.cell(row=22, column=9).value or 0)

    kst2 = 0.0
    if (ws_kot2.cell(row=16, column=10).value or 0) >= 0.1:
        qkotn2 = (ws_kot2.cell(row=16, column=10).value or 0) / (
            (ws_kot2.cell(row=17, column=5).value or 0) + (ws_kot2.cell(row=17, column=6).value or 0) +
            (ws_kot2.cell(row=17, column=7).value or 0) + (ws_kot2.cell(row=17, column=8).value or 0) +
            (ws_kot2.cell(row=17, column=9).value or 0)
        ) / 129.4

        if qkotn2 >= 0.75:
            x1 = 15.05 - 29.4388 * qkotn2 + 14.3971 * qkotn2 * qkotn2
            x2 = 27.43 - 54.3532 * qkotn2 + 26.9242 * qkotn2 * qkotn2
            kst2u = x1 + (x2 - x1) / 0.25 * (qkotn2 - 0.75)
        else:
            x1 = 11.22 - 21.9621 * qkotn2 + 10.7443 * qkotn2 * qkotn2
            x2 = 15.05 - 29.4388 * qkotn2 + 14.3971 * qkotn2 * qkotn2
            kst2u = x1 + (x2 - x1) / 0.25 * (qkotn2 - 0.5)

        if qkotn2 >= 0.75:
            x1 = 1.78 - 3.5413 * qkotn2 + 1.7596 * qkotn2 * qkotn2
            x2 = 2.24 - 4.4073 * qkotn2 + 2.1738 * qkotn2 * qkotn2
            kst2g = x1 + (x2 - x1) / 0.25 * (qkotn2 - 0.75)
        else:
            x1 = 1.22 - 2.4659 * qkotn2 + 1.2484 * qkotn2 * qkotn2
            x2 = 1.78 - 3.5413 * qkotn2 + 1.7596 * qkotn2 * qkotn2
            kst2g = x1 + (x2 - x1) / 0.25 * (qkotn2 - 0.5)

        kst2 = kst2u * (ws_kot2.cell(row=21, column=10).value or 0) + kst2g * (ws_kot2.cell(row=22, column=10).value or 0)

    ws_tep.cell(row=163, column=5).value = (
        kst1 * (ws_kot1.cell(row=15, column=9).value or 0) + kst2 * (ws_kot2.cell(row=15, column=10).value or 0)
    ) / ((ws_kot1.cell(row=15, column=9).value or 0) + (ws_kot2.cell(row=15, column=10).value or 0))

    ws_tep.cell(row=165, column=5).value = 6
    ws_tep.cell(row=164, column=5).value = (
        (ws_tep.cell(row=159, column=5).value or 0) *
        (ws_tep.cell(row=160, column=5).value or 0) *
        (1 + (ws_tep.cell(row=163, column=5).value or 0) / 100) /
        7 / (ws_tep.cell(row=157, column=5).value or 1) /
        (ws_tep.cell(row=158, column=5).value or 1) /
        (ws_tep.cell(row=161, column=5).value or 1) * 10000
    )
    ws_tep.cell(row=166, column=5).value = (ws_tep.cell(row=164, column=5).value or 0) * (1 + (ws_tep.cell(row=165, column=5).value or 0) / 100)

    # ===================== ФІНАЛЬНИЙ РЯДОК 167 =====================
    if (ws_tep.cell(row=14, column=5).value or 0) > 0:
        ws_tep.cell(row=167, column=5).value = (
            (100 - (ws_tep.cell(row=25, column=5).value or 0) - (ws_tep.cell(row=14, column=5).value or 0) + (ws_tep.cell(row=84, column=5).value or 0)) *
            (1 + (ws_tep.cell(row=163, column=5).value or 0) / 100) * 100000 /
            7 / (ws_tep.cell(row=157, column=5).value or 1) /
            (ws_tep.cell(row=158, column=5).value or 1) /
            (ws_tep.cell(row=161, column=5).value or 1) +
            (ws_tep.cell(row=14, column=5).value or 0) * 1000 / 7 / (ws_pwk.cell(row=71, column=10).value or 1) +
            ((ws_tep.cell(row=69, column=5).value or 0) + (ws_tep.cell(row=75, column=5).value or 0)) /
            (ws_tep.cell(row=9, column=5).value or 1) * (ws_tep.cell(row=148, column=5).value or 0) * (ws_tep.cell(row=166, column=5).value or 0)
        )
    else:
        ws_tep.cell(row=167, column=5).value = (
            (100 - (ws_tep.cell(row=25, column=5).value or 0) - (ws_tep.cell(row=14, column=5).value or 0) + (ws_tep.cell(row=84, column=5).value or 0)) *
            (1 + (ws_tep.cell(row=163, column=5).value or 0) / 100) * 100000 /
            7 / (ws_tep.cell(row=157, column=5).value or 1) /
            (ws_tep.cell(row=158, column=5).value or 1) /
            (ws_tep.cell(row=161, column=5).value or 1) +
            ((ws_tep.cell(row=69, column=5).value or 0) + (ws_tep.cell(row=75, column=5).value or 0)) /
            (ws_tep.cell(row=9, column=5).value or 1) * (ws_tep.cell(row=148, column=5).value or 0) * (ws_tep.cell(row=166, column=5).value or 0)
        )

    ws_tep.cell(row=168, column=5).value = 3
    ws_tep.cell(row=169, column=5).value = (ws_tep.cell(row=167, column=5).value or 0) * (1 + (ws_tep.cell(row=168, column=5).value or 0) / 100)

    ws_tep.cell(row=170, column=5).value = (
        (ws_kot1.cell(row=14, column=9).value or 0) +
        (ws_kot2.cell(row=14, column=10).value or 0) +
        (ws_pwk.cell(row=15, column=10).value or 0) -
        ((ws_tep.cell(row=166, column=5).value or 0) * (ws_tep.cell(row=88, column=5).value or 0) +
         (ws_tep.cell(row=169, column=5).value or 0) * (ws_tep.cell(row=9, column=5).value or 0)) / 1000
    )


# =============================================
# ПРИКЛАД ВИКОРИСТАННЯ
# =============================================
if __name__ == "__main__":
    wb = load_workbook("/data/exel/cerkassy_test_ОБНУЛЕНО_20260402_144235.xlsx")          # ← твій Excel-файл

    ws_tep = wb["ТЕП"]
    ws_tur1 = wb["Турбіна I черга"]
    ws_tur2 = wb["Турбіна II черга"]
    ws_kot1 = wb["Котел I черга"]
    ws_kot2 = wb["Котел II черга"]
    ws_pwk = wb["ПВК"]

    nmesac = 5                                    # ← номер місяця (1-12)

    calc_teps(ws_tep, ws_tur1, ws_tur2, ws_kot1, ws_kot2, ws_pwk, nmesac)

    wb.save("G:\\other\\PTV\\Py_calculation\\data\\exel\\cerkassy_test_teps_calc.xlsx")
    print("✅ Розрахунок TEP завершено! Файл збережено як your_file_updated.xlsx")