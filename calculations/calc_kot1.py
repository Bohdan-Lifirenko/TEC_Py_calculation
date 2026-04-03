import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import os
from datetime import datetime

# ====================== ВСІ ФУНКЦІЇ З ПОПЕРЕДНЬОГО КОДУ ======================

def get_cell(sheet: Worksheet, row: int, col: int):
    cell_value = sheet.cell(row=row, column=col).value
    return 0 if cell_value is None else float(cell_value)

def set_cell(sheet: Worksheet, row: int, col: int, value):
    sheet.cell(row=row, column=col).value = value

def calculate_tpw(sht_tur1: Worksheet, sht_tur2: Worksheet):
    if get_cell(sht_tur1, 8, 5) == 0:
        tpw1 = 0
    else:
        val = get_cell(sht_tur1, 139, 5)
        tpw1 = 154.7 + 0.53667 * val - 0.00088 * val * val

    if get_cell(sht_tur1, 8, 6) == 0:
        tpw2 = 0
    else:
        val = get_cell(sht_tur1, 139, 6)
        tpw2 = 156.6 + 0.52364 * val - 0.00083 * val * val

    if get_cell(sht_tur2, 8, 5) == 0:
        tpw3 = 0
    else:
        val = get_cell(sht_tur2, 148, 5)
        tpw3 = 184 + 0.201104 * val - 0.0001689 * val * val

    if get_cell(sht_tur2, 8, 6) == 0:
        tpw4 = 0
    else:
        val = get_cell(sht_tur2, 148, 6)
        tpw4 = 184 + 0.201104 * val - 0.0001689 * val * val

    if get_cell(sht_tur2, 8, 7) == 0:
        tpw5 = 0
    else:
        val = get_cell(sht_tur2, 148, 7)
        tpw5 = 184 + 0.201104 * val - 0.0001689 * val * val

    return tpw1, tpw2, tpw3, tpw4, tpw5


def process_single_column(
    sht_kot1: Worksheet, col: int,
    tpw1: float, tpw2: float, tpw3: float, tpw4: float, tpw5: float,
    sht_tur1: Worksheet, sht_tur2: Worksheet
):
    if get_cell(sht_kot1, 17, col) == 0:
        for k in range(12, 26): set_cell(sht_kot1, k, col, 0)
        for k in range(28, 45): set_cell(sht_kot1, k, col, 0)
        for k in range(46, 69): set_cell(sht_kot1, k, col, 0)
        for k in range(70, 77): set_cell(sht_kot1, k, col, 0)
        for k in range(80, 113): set_cell(sht_kot1, k, col, 0)
        for k in range(118, 139): set_cell(sht_kot1, k, col, 0)
    else:
        set_cell(sht_kot1, 14, col, get_cell(sht_kot1, 12, col) + get_cell(sht_kot1, 13, col))
        set_cell(sht_kot1, 15, col, get_cell(sht_kot1, 16, col) / get_cell(sht_kot1, 17, col))
        set_cell(sht_kot1, 19, col, 254500)
        set_cell(sht_kot1, 20, col, get_cell(sht_kot1, 18, col) - get_cell(sht_kot1, 19, col))
        set_cell(sht_kot1, 22, col, get_cell(sht_kot1, 13, col) / get_cell(sht_kot1, 14, col))
        set_cell(sht_kot1, 21, col, 1 - get_cell(sht_kot1, 22, col))

        num = (tpw1 * get_cell(sht_tur1, 139, 5) + tpw2 * get_cell(sht_tur1, 139, 6) +
               tpw3 * get_cell(sht_tur2, 148, 5) + tpw4 * get_cell(sht_tur2, 148, 6) +
               tpw5 * get_cell(sht_tur2, 148, 7))
        den = (get_cell(sht_tur1, 139, 5) + get_cell(sht_tur1, 139, 6) +
               get_cell(sht_tur2, 148, 5) + get_cell(sht_tur2, 148, 6) +
               get_cell(sht_tur2, 148, 7))
        set_cell(sht_kot1, 23, col, num / den if den != 0 else 0)

        set_cell(sht_kot1, 28, col, get_cell(sht_kot1, 27, 9) + 4)

        # Блок Cells(21, col)
        c21 = get_cell(sht_kot1, 21, col)
        if c21 == 0:
            for k in [29,31,33,34,37,38,46,48,53,55,60,63,65,70,72,74,80,83,85,90,92,93,98,100,101,103,118,120,126,128,131,133]:
                set_cell(sht_kot1, k, col, 0)
        else:
            set_cell(sht_kot1, 34, col, 60)
            set_cell(sht_kot1, 37, col, get_cell(sht_kot1, 31, col) - get_cell(sht_kot1, 29, col))
            set_cell(sht_kot1, 38, col, 47)
            c15 = get_cell(sht_kot1, 15, col)
            set_cell(sht_kot1, 46, col, 2.1397 - 0.0254*c15 + 0.00019046*c15*c15)
            set_cell(sht_kot1, 48, col, get_cell(sht_kot1, 46, col)*(1 + get_cell(sht_kot1, 47, col)/100))
            set_cell(sht_kot1, 53, col, 0.2104 - 0.00242972*c15 + 0.0000122438*c15*c15)
            set_cell(sht_kot1, 55, col, get_cell(sht_kot1, 53, col)*(1 + get_cell(sht_kot1, 54, col)/100))
            set_cell(sht_kot1, 60, col, get_cell(sht_kot1, 48, col) + get_cell(sht_kot1, 55, col))
            set_cell(sht_kot1, 63, col, 137.9798 + 0.639*c15)
            set_cell(sht_kot1, 65, col, get_cell(sht_kot1, 63, col) + get_cell(sht_kot1, 64, col))
            set_cell(sht_kot1, 70, col, (get_cell(sht_kot1, 33, col)-60)*0.6)
            set_cell(sht_kot1, 72, col, (get_cell(sht_kot1, 37, col)-30)*0.3)
            set_cell(sht_kot1, 73, col, (get_cell(sht_kot1, 23, col)-226)*0.2)
            set_cell(sht_kot1, 74, col, get_cell(sht_kot1, 65, col) + get_cell(sht_kot1, 70, col) + get_cell(sht_kot1, 72, col) + get_cell(sht_kot1, 73, col))

            set_cell(sht_kot1, 90, col, 35.4 - 1.107*c15 + 0.00990781*c15*c15)
            set_cell(sht_kot1, 92, col, get_cell(sht_kot1, 90, col)*(1 + get_cell(sht_kot1, 91, col)/100))
            set_cell(sht_kot1, 93, col, 0.8 * get_cell(sht_kot1, 92, col) / (100 - get_cell(sht_kot1, 92, col)) * 7800 *
                                       get_cell(sht_kot1, 8, 9) / get_cell(sht_kot1, 7, 9))

            c77_9 = get_cell(sht_kot1, 77, 9)
            c78_9 = get_cell(sht_kot1, 78, 9)
            c79_9 = get_cell(sht_kot1, 79, 9)
            c60 = get_cell(sht_kot1, 60, col)
            c28 = get_cell(sht_kot1, 28, col)
            c74 = get_cell(sht_kot1, 74, col)
            c93 = get_cell(sht_kot1, 93, col)
            term1 = c77_9 * c60 + c78_9
            term2 = c74 - c60 * c28 / (c60 + c79_9)
            term3 = 0.9805 + 1.3 * c74 / 10000
            term4 = (1 - 0.01 * c93)
            term5 = 0.2 * 0.8 * get_cell(sht_kot1, 8, 9) * c74 / get_cell(sht_kot1, 7, 9)
            set_cell(sht_kot1, 80, col, term1 * term2 * term3 * term4 / 100 + term5)

            set_cell(sht_kot1, 83, col, 0)
            set_cell(sht_kot1, 85, col, get_cell(sht_kot1, 83, col)*(1 + get_cell(sht_kot1, 84, col)/100))

            set_cell(sht_kot1, 95, col, 3.4984 - 0.1059*c15 + 0.00139459*c15*c15 - 0.0000067467*c15*c15*c15)
            set_cell(sht_kot1, 97, col, get_cell(sht_kot1, 95, col)*(1 + get_cell(sht_kot1, 96, col)/100))
            set_cell(sht_kot1, 98, col, 0.2 * 1400 * get_cell(sht_kot1, 8, 9) / get_cell(sht_kot1, 7, 9))
            set_cell(sht_kot1, 100, col, get_cell(sht_kot1, 98, col)*(1 + get_cell(sht_kot1, 99, col)/100))
            set_cell(sht_kot1, 101, col, 0.4)
            set_cell(sht_kot1, 103, col, get_cell(sht_kot1, 101, col)*(1 + get_cell(sht_kot1, 102, col)/100))
            set_cell(sht_kot1, 118, col, 10.7518 - 0.0382*c15)
            set_cell(sht_kot1, 120, col, get_cell(sht_kot1, 118, col)*(1 + get_cell(sht_kot1, 119, col)/100))
            set_cell(sht_kot1, 126, col, 34.1415 - 0.1437*c15)
            set_cell(sht_kot1, 128, col, get_cell(sht_kot1, 126, col)*(1 + get_cell(sht_kot1, 127, col)/100))
            set_cell(sht_kot1, 131, col, 0.3843 + 0.00162226*c15 - 0.000000308006*c15*c15)
            set_cell(sht_kot1, 133, col, get_cell(sht_kot1, 131, col)*(1 + get_cell(sht_kot1, 132, col)/100))

        # Блок Cells(22, col)
        c22 = get_cell(sht_kot1, 22, col)
        if c22 == 0:
            for k in [30,32,35,36,39,40,49,51,56,58,61,66,68,71,75,81,86,88,112,121,123,134,136]:
                set_cell(sht_kot1, k, col, 0)
        else:
            set_cell(sht_kot1, 36, col, 30)
            set_cell(sht_kot1, 39, col, get_cell(sht_kot1, 32, col) - get_cell(sht_kot1, 30, col))
            set_cell(sht_kot1, 40, col, 17)
            c15 = get_cell(sht_kot1, 15, col)
            set_cell(sht_kot1, 49, col, 2.0489 - 0.0222*c15 + 0.000136328*c15*c15)
            set_cell(sht_kot1, 51, col, get_cell(sht_kot1, 49, col)*(1 + get_cell(sht_kot1, 50, col)/100))
            set_cell(sht_kot1, 56, col, 0.2011 - 0.00211662*c15 + 0.00000966484*c15*c15)
            set_cell(sht_kot1, 58, col, get_cell(sht_kot1, 56, col)*(1 + get_cell(sht_kot1, 57, col)/100))
            set_cell(sht_kot1, 61, col, get_cell(sht_kot1, 51, col) + get_cell(sht_kot1, 58, col))
            set_cell(sht_kot1, 66, col, 106.2661 + 0.5997*c15)
            set_cell(sht_kot1, 68, col, get_cell(sht_kot1, 66, col) + get_cell(sht_kot1, 67, col))
            set_cell(sht_kot1, 71, col, (get_cell(sht_kot1, 35, col)-30)*0.6)
            set_cell(sht_kot1, 73, col, (get_cell(sht_kot1, 23, col)-226)*0.2)
            set_cell(sht_kot1, 75, col, get_cell(sht_kot1, 68, col) + get_cell(sht_kot1, 71, col) + get_cell(sht_kot1, 73, col))

            c61 = get_cell(sht_kot1, 61, col)
            c28 = get_cell(sht_kot1, 28, col)
            c75 = get_cell(sht_kot1, 75, col)
            set_cell(sht_kot1, 81, col, (3.53*c61 + 0.6) * (c75 - c61*c28/(c61 + 0.18)) * (0.9805 + 1.3*c75/10000) / 100)

            set_cell(sht_kot1, 86, col, 0.03)
            set_cell(sht_kot1, 88, col, get_cell(sht_kot1, 86, col)*(1 + get_cell(sht_kot1, 87, col)/100))

            set_cell(sht_kot1, 95, col, 3.4984 - 0.1059*c15 + 0.00139459*c15*c15 - 0.0000067467*c15*c15*c15)
            set_cell(sht_kot1, 97, col, get_cell(sht_kot1, 95, col)*(1 + get_cell(sht_kot1, 96, col)/100))
            set_cell(sht_kot1, 121, col, 16.7181 - 0.2858*c15 + 0.00181036*c15*c15)
            set_cell(sht_kot1, 123, col, get_cell(sht_kot1, 121, col)*(1 + get_cell(sht_kot1, 122, col)/100))
            set_cell(sht_kot1, 134, col, 0.3305 + 0.00155091*c15 + 0.000000198932*c15*c15)
            set_cell(sht_kot1, 136, col, get_cell(sht_kot1, 134, col)*(1 + get_cell(sht_kot1, 135, col)/100))

        # Загальні розрахунки
        set_cell(sht_kot1, 42, col, get_cell(sht_kot1, 41, col) * 17.5)
        set_cell(sht_kot1, 43, col, get_cell(sht_kot1, 41, col) * 2.8)
        set_cell(sht_kot1, 44, col, get_cell(sht_kot1, 41, col) * 14.9)

        set_cell(sht_kot1, 52, col, get_cell(sht_kot1, 48, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 51, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 59, col, get_cell(sht_kot1, 55, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 58, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 62, col, get_cell(sht_kot1, 60, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 61, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 76, col, get_cell(sht_kot1, 74, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 75, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 82, col, get_cell(sht_kot1, 80, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 81, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 89, col, get_cell(sht_kot1, 85, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 88, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 94, col, get_cell(sht_kot1, 93, col)*get_cell(sht_kot1, 21, col))
        set_cell(sht_kot1, 104, col, (get_cell(sht_kot1, 100, col) + get_cell(sht_kot1, 103, col)) * get_cell(sht_kot1, 21, col))
        set_cell(sht_kot1, 106, col, 0.008 * get_cell(sht_kot1, 20, col) / 100000 * 100)
        set_cell(sht_kot1, 108, col, get_cell(sht_kot1, 107, col) / 100 * get_cell(sht_kot1, 82, col))

        set_cell(sht_kot1, 109, col, 100 - (get_cell(sht_kot1, 82, col) + get_cell(sht_kot1, 89, col) +
                                            get_cell(sht_kot1, 94, col) + get_cell(sht_kot1, 97, col) +
                                            get_cell(sht_kot1, 104, col) + get_cell(sht_kot1, 106, col) +
                                            get_cell(sht_kot1, 108, col)))

        set_cell(sht_kot1, 110, col, get_cell(sht_kot1, 16, col)*100/7/get_cell(sht_kot1, 109, col) + get_cell(sht_kot1, 42, col))

        c7_9 = get_cell(sht_kot1, 7, 9)
        if c7_9 == 0:
            set_cell(sht_kot1, 111, col, 0)
        else:
            set_cell(sht_kot1, 111, col, get_cell(sht_kot1, 110, col)*7000/c7_9*get_cell(sht_kot1, 21, col))

        c10_9 = get_cell(sht_kot1, 10, 9)
        if c10_9 == 0:
            set_cell(sht_kot1, 112, col, 0)
        else:
            set_cell(sht_kot1, 112, col, get_cell(sht_kot1, 110, col)*7000/c10_9*get_cell(sht_kot1, 22, col))

        set_cell(sht_kot1, 124, col, get_cell(sht_kot1, 120, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 123, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 125, col, get_cell(sht_kot1, 124, col)*get_cell(sht_kot1, 16, col)/1000)
        set_cell(sht_kot1, 129, col, get_cell(sht_kot1, 128, col)*get_cell(sht_kot1, 111, col)/1000)
        set_cell(sht_kot1, 130, col, get_cell(sht_kot1, 125, col) + get_cell(sht_kot1, 129, col) + get_cell(sht_kot1, 43, col))
        set_cell(sht_kot1, 137, col, get_cell(sht_kot1, 133, col)*get_cell(sht_kot1, 21, col) + get_cell(sht_kot1, 136, col)*get_cell(sht_kot1, 22, col))
        set_cell(sht_kot1, 138, col, get_cell(sht_kot1, 137, col)*get_cell(sht_kot1, 17, col) + get_cell(sht_kot1, 44, col))


def calc_kot1s(sht_kot1: Worksheet, sht_tur1: Worksheet, sht_tur2: Worksheet):
    c7_9 = get_cell(sht_kot1, 7, 9)
    if c7_9 == 0:
        set_cell(sht_kot1, 77, 9, 0)
        set_cell(sht_kot1, 78, 9, 0)
        set_cell(sht_kot1, 79, 9, 0)
    else:
        ratio = get_cell(sht_kot1, 9, 9) / c7_9 * 1000
        set_cell(sht_kot1, 77, 9, 3.5 + 0.02 * ratio)
        set_cell(sht_kot1, 78, 9, 0.4 + 0.04 * ratio)
        set_cell(sht_kot1, 79, 9, 0.14 if ratio < 2 else 0.12 + 0.014 * ratio)

    tpw1, tpw2, tpw3, tpw4, tpw5 = calculate_tpw(sht_tur1, sht_tur2)
    process_single_column(sht_kot1, 5, tpw1, tpw2, tpw3, tpw4, tpw5, sht_tur1, sht_tur2)
    process_single_column(sht_kot1, 6, tpw1, tpw2, tpw3, tpw4, tpw5, sht_tur1, sht_tur2)


def calc_kot1m(sht_kot1: Worksheet, sht_tur1: Worksheet, sht_tur2: Worksheet):
    tpw1, tpw2, tpw3, tpw4, tpw5 = calculate_tpw(sht_tur1, sht_tur2)
    process_single_column(sht_kot1, 7, tpw1, tpw2, tpw3, tpw4, tpw5, sht_tur1, sht_tur2)
    process_single_column(sht_kot1, 8, tpw1, tpw2, tpw3, tpw4, tpw5, sht_tur1, sht_tur2)

    set_cell(sht_kot1, 110, 9, get_cell(sht_kot1, 110, 5) + get_cell(sht_kot1, 110, 6) + get_cell(sht_kot1, 110, 7) + get_cell(sht_kot1, 110, 8))
    set_cell(sht_kot1, 111, 9, get_cell(sht_kot1, 111, 5) + get_cell(sht_kot1, 111, 6) + get_cell(sht_kot1, 111, 7) + get_cell(sht_kot1, 111, 8))
    set_cell(sht_kot1, 112, 9, get_cell(sht_kot1, 112, 5) + get_cell(sht_kot1, 112, 6) + get_cell(sht_kot1, 112, 7) + get_cell(sht_kot1, 112, 8))

    c110_9 = get_cell(sht_kot1, 110, 9)
    if c110_9 == 0:
        for k in [113,114,115,116,21,22,15,125,129,130,138]:
            set_cell(sht_kot1, k, 9, 0)
    else:
        set_cell(sht_kot1, 12, 9, get_cell(sht_kot1, 12, 5)+get_cell(sht_kot1, 12, 6)+get_cell(sht_kot1, 12, 7)+get_cell(sht_kot1, 12, 8))
        set_cell(sht_kot1, 13, 9, get_cell(sht_kot1, 13, 5)+get_cell(sht_kot1, 13, 6)+get_cell(sht_kot1, 13, 7)+get_cell(sht_kot1, 13, 8))
        set_cell(sht_kot1, 14, 9, get_cell(sht_kot1, 14, 5)+get_cell(sht_kot1, 14, 6)+get_cell(sht_kot1, 14, 7)+get_cell(sht_kot1, 14, 8))
        set_cell(sht_kot1, 16, 9, get_cell(sht_kot1, 16, 5)+get_cell(sht_kot1, 16, 6)+get_cell(sht_kot1, 16, 7)+get_cell(sht_kot1, 16, 8))

        sum17 = (get_cell(sht_kot1, 17, 5) + get_cell(sht_kot1, 17, 6) + get_cell(sht_kot1, 17, 7) + get_cell(sht_kot1, 17, 8))
        set_cell(sht_kot1, 15, 9, 0 if sum17 == 0 else get_cell(sht_kot1, 16, 9) / sum17)
        set_cell(sht_kot1, 17, 9, sum17)

        set_cell(sht_kot1, 113, 9, (get_cell(sht_kot1, 109, 5)*get_cell(sht_kot1, 110, 5) +
                                    get_cell(sht_kot1, 109, 6)*get_cell(sht_kot1, 110, 6) +
                                    get_cell(sht_kot1, 109, 7)*get_cell(sht_kot1, 110, 7) +
                                    get_cell(sht_kot1, 109, 8)*get_cell(sht_kot1, 110, 8)) / c110_9)
        set_cell(sht_kot1, 114, 9, (get_cell(sht_kot1, 76, 5)*get_cell(sht_kot1, 110, 5) +
                                    get_cell(sht_kot1, 76, 6)*get_cell(sht_kot1, 110, 6) +
                                    get_cell(sht_kot1, 76, 7)*get_cell(sht_kot1, 110, 7) +
                                    get_cell(sht_kot1, 76, 8)*get_cell(sht_kot1, 110, 8)) / c110_9)
        set_cell(sht_kot1, 115, 9, (get_cell(sht_kot1, 52, 5)*get_cell(sht_kot1, 110, 5) +
                                    get_cell(sht_kot1, 52, 6)*get_cell(sht_kot1, 110, 6) +
                                    get_cell(sht_kot1, 52, 7)*get_cell(sht_kot1, 110, 7) +
                                    get_cell(sht_kot1, 52, 8)*get_cell(sht_kot1, 110, 8)) / c110_9)
        set_cell(sht_kot1, 116, 9, (get_cell(sht_kot1, 62, 5)*get_cell(sht_kot1, 110, 5) +
                                    get_cell(sht_kot1, 62, 6)*get_cell(sht_kot1, 110, 6) +
                                    get_cell(sht_kot1, 62, 7)*get_cell(sht_kot1, 110, 7) +
                                    get_cell(sht_kot1, 62, 8)*get_cell(sht_kot1, 110, 8)) / c110_9)

        set_cell(sht_kot1, 21, 9, get_cell(sht_kot1, 111, 9) * get_cell(sht_kot1, 7, 9) / c110_9 / 7000)
        set_cell(sht_kot1, 22, 9, get_cell(sht_kot1, 112, 9) * get_cell(sht_kot1, 10, 9) / c110_9 / 7000)

        set_cell(sht_kot1, 125, 9, get_cell(sht_kot1, 125, 5)+get_cell(sht_kot1, 125, 6)+get_cell(sht_kot1, 125, 7)+get_cell(sht_kot1, 125, 8))
        set_cell(sht_kot1, 129, 9, get_cell(sht_kot1, 129, 5)+get_cell(sht_kot1, 129, 6)+get_cell(sht_kot1, 129, 7)+get_cell(sht_kot1, 129, 8))
        set_cell(sht_kot1, 130, 9, get_cell(sht_kot1, 130, 5)+get_cell(sht_kot1, 130, 6)+get_cell(sht_kot1, 130, 7)+get_cell(sht_kot1, 130, 8))
        set_cell(sht_kot1, 138, 9, get_cell(sht_kot1, 138, 5)+get_cell(sht_kot1, 138, 6)+get_cell(sht_kot1, 138, 7)+get_cell(sht_kot1, 138, 8))


# ====================== ГРАФІЧНИЙ ІНТЕРФЕЙС ======================

class KotCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Калькулятор котлів (VBA → Python)")
        self.root.geometry("700x400")
        self.root.resizable(False, False)

        self.file_path = tk.StringVar()

        # Стиль
        style = ttk.Style()
        style.theme_use('clam')

        # Заголовок
        tk.Label(root, text="Десктопний калькулятор котлів", font=("Arial", 16, "bold")).pack(pady=15)

        # Фрейм вибору файлу
        file_frame = tk.Frame(root)
        file_frame.pack(pady=10, padx=20, fill="x")

        tk.Label(file_frame, text="Excel файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(file_frame, textvariable=self.file_path, width=60, state="readonly").pack(side="left", padx=10)
        tk.Button(file_frame, text="Обрати файл", command=self.select_file, width=15).pack(side="left")

        # Інформація про листи
        info = tk.Label(root, text="Очікувані назви листів:\nKot1  |  Tur1  |  Tur2\n(якщо назви інші — програма не спрацює)",
                        font=("Arial", 9), fg="gray", justify="center")
        info.pack(pady=5)

        # Кнопка розрахунку
        self.calc_btn = tk.Button(root, text="ВИКОНАТИ РОЗРАХУНОК", font=("Arial", 12, "bold"),
                                  bg="#007ACC", fg="white", height=2, command=self.run_calculation)
        self.calc_btn.pack(pady=20)

        # Статус
        self.status = tk.Label(root, text="Готовий до роботи", fg="green", font=("Arial", 10))
        self.status.pack(pady=10)

        # Лог
        self.log_text = tk.Text(root, height=8, width=80, state="disabled")
        self.log_text.pack(pady=10, padx=20)

    def log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def select_file(self):
        path = filedialog.askopenfilename(
            title="Виберіть Excel файл (.xlsx)",
            filetypes=[("Excel файли", "*.xlsx *.xlsm")]
        )
        if path:
            self.file_path.set(path)
            self.log(f"Обрано файл: {os.path.basename(path)}")

    def run_calculation(self):
        if not self.file_path.get():
            messagebox.showerror("Помилка", "Спочатку оберіть Excel файл!")
            return

        self.calc_btn.config(state="disabled")
        self.status.config(text="Обчислення...", fg="orange")
        self.log("Початок розрахунку...")

        try:
            wb = openpyxl.load_workbook(self.file_path.get(), data_only=True)

            # Назви листів (змінюйте тут, якщо у вас інші назви)
            sht_kot1 = wb["Котел I черга"]
            sht_tur1 = wb["Турбіна I черга"]
            sht_tur2 = wb["Турбіна II черга"]

            self.log("Листа завантажено успішно")

            calc_kot1s(sht_kot1, sht_tur1, sht_tur2)
            self.log("Виконано CalcKot1s")

            calc_kot1m(sht_kot1, sht_tur1, sht_tur2)
            self.log("Виконано CalcKot1m")

            # Зберігаємо з новою назвою
            base, ext = os.path.splitext(self.file_path.get())
            new_path = f"{base}_РАЗРАХОВАНО_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            wb.save(new_path)

            self.log(f"Файл збережено: {os.path.basename(new_path)}")
            self.status.config(text="Розрахунок завершено успішно!", fg="green")

            messagebox.showinfo("Готово!",
                f"Розрахунок виконано!\n\nЗбережено як:\n{os.path.basename(new_path)}")

        except FileNotFoundError:
            messagebox.showerror("Помилка", "Файл не знайдено!")
            self.status.config(text="Помилка", fg="red")
        except KeyError as e:
            messagebox.showerror("Помилка", f"Не знайдено лист: {e}\nПеревірте назви листів (Kot1, Tur1, Tur2)")
            self.status.config(text="Помилка", fg="red")
        except Exception as e:
            messagebox.showerror("Помилка", f"Сталася помилка:\n{str(e)}")
            self.status.config(text="Помилка", fg="red")
            self.log(f"ПОМИЛКА: {e}")
        finally:
            self.calc_btn.config(state="normal")


if __name__ == "__main__":
    root = tk.Tk()
    app = KotCalculatorApp(root)
    root.mainloop()