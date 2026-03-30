from openpyxl import load_workbook

wb = load_workbook("Book1.xlsm")

print(wb.sheetnames)

wb = load_workbook("Book1.xlsm")

sht_kot1 = wb["Котел I черга"]

for row in sht_kot1.iter_rows(values_only=True):
    print(row)