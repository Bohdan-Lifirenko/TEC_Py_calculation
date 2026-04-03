# gui/kot_calculator_app.py
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os
from datetime import datetime

from calculations.calc_kot1 import calc_kot1s, calc_kot1m
from json_saving import save_excel_state, load_json_to_excel
from .base_tab import BaseTab


class KotCalculatorApp(BaseTab):
    def __init__(self, parent_frame):
        super().__init__(parent_frame)

        tk.Label(self.frame, text="Десктопний калькулятор котлів", font=("Arial", 16, "bold")).pack(pady=15)

        file_frame = tk.Frame(self.frame)
        file_frame.pack(pady=10, padx=20, fill="x")
        tk.Label(file_frame, text="Excel файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(file_frame, textvariable=self.file_path, width=60, state="readonly").pack(side="left", padx=10)
        tk.Button(file_frame, text="Обрати файл", command=self.select_file, width=15).pack(side="left")

        info = tk.Label(self.frame, text="Очікувані назви листів:\nКотел I черга | Турбіна I черга | Турбіна II черга",
                        font=("Arial", 9), fg="gray", justify="center")
        info.pack(pady=5)

        self.calc_btn = tk.Button(self.frame, text="ВИКОНАТИ РОЗРАХУНОК", font=("Arial", 12, "bold"),
                                  bg="#007ACC", fg="white", height=2, command=self.run_calculation)
        self.calc_btn.pack(pady=20)

        # Кнопки сейвів
        btn_frame = tk.Frame(self.frame)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="💾 Зберегти сейв (JSON)",
                  command=self.save_current_state, width=20).pack(side="left", padx=5)
        tk.Button(btn_frame, text="📂 Завантажити сейв",
                  command=self.load_save_file, width=20).pack(side="left", padx=5)

        self.status = tk.Label(self.frame, text="Готовий до роботи", fg="green", font=("Arial", 10))
        self.status.pack(pady=10)

        self.create_log_widget(height=10)

    def run_calculation(self):
        if not self.file_path.get():
            messagebox.showerror("Помилка", "Оберіть Excel файл!")
            return

        self.calc_btn.config(state="disabled")
        self.status.config(text="Обчислення...", fg="orange")
        self.log("Початок розрахунку...")

        try:
            wb = openpyxl.load_workbook(self.file_path.get(), data_only=True)
            sht_kot1 = wb["Котел I черга"]
            sht_tur1 = wb["Турбіна I черга"]
            sht_tur2 = wb["Турбіна II черга"]

            self.log("Листа завантажено успішно")
            calc_kot1s(sht_kot1, sht_tur1, sht_tur2)
            self.log("Виконано CalcKot1s")
            calc_kot1m(sht_kot1, sht_tur1, sht_tur2)
            self.log("Виконано CalcKot1m")

            base, ext = os.path.splitext(self.file_path.get())
            new_path = f"{base}_РАЗРАХОВАНО_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            wb.save(new_path)

            self.log(f"Файл збережено: {os.path.basename(new_path)}")
            self.status.config(text="Розрахунок завершено успішно!", fg="green")
            messagebox.showinfo("Готово!", f"Збережено як:\n{os.path.basename(new_path)}")
        except Exception as e:
            self.log(f"ПОМИЛКА: {e}")
            messagebox.showerror("Помилка", str(e))
            self.status.config(text="Помилка", fg="red")
        finally:
            self.calc_btn.config(state="normal")

    def save_current_state(self):
        if not self.file_path.get():
            messagebox.showerror("Помилка", "Спочатку оберіть Excel файл!")
            return
        wb = openpyxl.load_workbook(self.file_path.get())
        base = os.path.splitext(self.file_path.get())[0]
        save_path = f"{base}_SEIV_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        save_excel_state(wb, save_path)
        messagebox.showinfo("Готово", f"Сейв збережено:\n{save_path}")

    def load_save_file(self):
        path = filedialog.askopenfilename(
            title="Виберіть сейв-файл",
            filetypes=[("JSON сейви", "*.json")]
        )
        if path:
            wb = openpyxl.load_workbook(self.file_path.get())
            load_json_to_excel(path, wb)
            wb.save(self.file_path.get())
            messagebox.showinfo("Готово", "Дані з сейву завантажено в Excel!")