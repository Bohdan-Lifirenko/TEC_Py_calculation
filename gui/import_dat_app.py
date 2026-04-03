# gui/import_dat_app.py
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

from .base_tab import BaseTab


class ImportDatApp(BaseTab):
    def __init__(self, parent_frame):
        super().__init__(parent_frame)
        self.dat_path = tk.StringVar()
        self.xlsx_path = tk.StringVar()

        tk.Label(self.frame, text="Імпорт даних з .dat у Excel",
                 font=("Arial", 16, "bold")).pack(pady=15)

        # .dat файл
        frame_dat = tk.Frame(self.frame)
        frame_dat.pack(pady=8, padx=20, fill="x")
        tk.Label(frame_dat, text=".dat файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(frame_dat, textvariable=self.dat_path, width=50, state="readonly").pack(side="left", padx=10)
        tk.Button(frame_dat, text="Обрати .dat", command=self.select_dat).pack(side="left")

        # Excel файл
        frame_xlsx = tk.Frame(self.frame)
        frame_xlsx.pack(pady=8, padx=20, fill="x")
        tk.Label(frame_xlsx, text="Excel файл:", font=("Arial", 11)).pack(side="left")
        tk.Entry(frame_xlsx, textvariable=self.xlsx_path, width=50, state="readonly").pack(side="left", padx=10)
        tk.Button(frame_xlsx, text="Обрати .xlsx", command=self.select_xlsx).pack(side="left")

        # Кнопка імпорту
        tk.Button(self.frame, text="ВИКОНАТИ ІМПОРТ .dat → Excel",
                  font=("Arial", 12, "bold"), bg="#28a745", fg="white", height=2,
                  command=self.run_import).pack(pady=20)

        self.create_log_widget(height=12)

    def select_dat(self):
        path = filedialog.askopenfilename(title="Виберіть .dat файл", filetypes=[("DAT файли", "*.dat")])
        if path:
            self.dat_path.set(path)
            self.log(f"Обрано .dat: {os.path.basename(path)}")

    def select_xlsx(self):
        path = filedialog.askopenfilename(title="Виберіть Excel файл", filetypes=[("Excel файли", "*.xlsx *.xlsm")])
        if path:
            self.xlsx_path.set(path)
            self.log(f"Обрано Excel: {os.path.basename(path)}")

    def run_import(self):
        dat = self.dat_path.get().strip()
        xlsx = self.xlsx_path.get().strip()
        if not dat or not xlsx:
            messagebox.showerror("Помилка", "Оберіть обидва файли!")
            return

        self.log("Початок імпорту...")
        try:
            import_dat_to_excel(dat, xlsx)
            self.log("✅ Імпорт завершено успішно!")
            messagebox.showinfo("Готово", f"Дані з .dat успішно імпортовано!\nЗбережено у:\n{os.path.basename(xlsx)}")
        except Exception as e:
            self.log(f"❌ ПОМИЛКА: {e}")
            messagebox.showerror("Помилка", str(e))


# ====================== ФУНКЦІЯ ІМПОРТУ ======================
def import_dat_to_excel(dat_path: str, xlsx_path: str):
    with open(dat_path, 'r', encoding='windows-1251') as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    if not lines:
        print("❌ .dat файл порожній!")
        return

    data_lines = lines[1:] if lines[0].startswith('strtDT') else lines

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    for line in data_lines:
        parts = line.split(';')
        if len(parts) < 4:
            continue

        sheet_name = parts[0].strip()
        try:
            row = int(parts[1].strip())
            col = int(parts[2].strip())
        except ValueError:
            continue

        value_str = parts[3].strip()
        if value_str == '-':
            cell_value = '-'
        else:
            try:
                cell_value = float(value_str.replace(',', '.')) if ',' in value_str else float(value_str)
            except ValueError:
                cell_value = value_str

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        ws.cell(row=row, column=col, value=cell_value)

    wb.save(xlsx_path)
    print(f"✅ Імпортовано {len(data_lines)} записів у {xlsx_path}")