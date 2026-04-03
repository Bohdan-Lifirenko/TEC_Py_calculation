import sys
import os
from openpyxl import load_workbook, Workbook

def import_dat_to_excel(dat_path: str, xlsx_path: str):
    # Читаємо .dat файл
    with open(dat_path, 'r', encoding='windows-1251') as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    if not lines:
        print("❌ .dat файл порожній!")
        return

    # Перший рядок — це дата (strtDT), просто пропускаємо
    data_lines = lines[1:] if lines[0].startswith('strtDT') else lines

    # Відкриваємо або створюємо Excel-файл
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        print(f"✅ Відкрито існуючий файл: {xlsx_path}")
    else:
        wb = Workbook()
        # Видаляємо пустий аркуш за замовчуванням
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        print(f"✅ Створено новий файл: {xlsx_path}")

    for line in data_lines:
        parts = line.split(';')
        print(parts)
        if len(parts) < 4:
            print(f"⚠️ Пропущено неправильний рядок: {line}")
            continue

        sheet_name = parts[0].strip()
        try:
            row = int(parts[1].strip())
            col = int(parts[2].strip())
        except ValueError:
            print(f"⚠️ Помилка в номерах рядка/стовпчика: {line}")
            continue

        value_str = parts[3].strip()

        # Автоматично визначаємо тип значення (число або текст)
        if value_str == '-':
            cell_value = '-'
        else:
            try:
                # Спробуємо перетворити на float (заміна коми на крапку)
                if ',' in value_str:
                    cell_value = float(value_str.replace(',', '.'))
                else:
                    cell_value = float(value_str) if '.' in value_str else int(value_str)
            except ValueError:
                cell_value = value_str  # залишаємо як текст

        # Отримуємо або створюємо аркуш
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # Записуємо значення в клітинку
        ws.cell(row=row, column=col, value=cell_value)

    # Зберігаємо файл
    wb.save(xlsx_path)
    print(f"✅ Успішно перенесено {len(data_lines)} значень у файл:")
    print(f"   📁 {xlsx_path}")


if __name__ == "__main__":
    import_dat_to_excel("data/saves/Март 2015.dat", "data/exel/cerkassy_test.xlsx")

    # print("=== Імпорт даних з .dat у Excel ===\n")
    #
    # if len(sys.argv) == 3:
    #     dat_file = sys.argv[1]
    #     xlsx_file = sys.argv[2]
    # else:
    #     dat_file = input("Введіть шлях до .dat файлу: ").strip()
    #     xlsx_file = input("Введіть шлях до .xlsx файлу (можна новий): ").strip()
    #
    # if not os.path.exists(dat_file):
    #     print(f"❌ Файл .dat не знайдено: {dat_file}")
    # else:
    #     import_dat_to_excel(dat_file, xlsx_file)