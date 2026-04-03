import json
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict

def save_excel_state(wb, save_path: str, sheets_to_save=None):
    """
    Зберігає стан вказаних аркушів у JSON-сейв
    sheets_to_save = ["Kot1", "Tur1", "Tur2"] або None (всі аркуші)
    """
    if sheets_to_save is None:
        sheets_to_save = wb.sheetnames

    data = {
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "original_file": getattr(wb, 'path', 'unknown.xlsx'),
        "sheets": {}
    }

    for sheet_name in sheets_to_save:
        print(sheet_name)
        if sheet_name not in wb:
            continue
        ws = wb[sheet_name]
        sheet_data = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:                    # зберігаємо тільки заповнені клітинки
                    key = f"{cell.row},{cell.column}"
                    sheet_data[key] = cell.value

        if sheet_data:                                        # якщо є дані
            data["sheets"][sheet_name] = sheet_data

    with open(save_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ Сейв створено: {save_path}")
    return save_path


def load_json_to_excel(json_path: str, wb):
    """Завантажує сейв назад у Excel"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    for sheet_name, cells in data.get("sheets", {}).items():
        if sheet_name not in wb:
            print(f"⚠️ Аркуш {sheet_name} не знайдено в файлі")
            continue
        ws = wb[sheet_name]

        for key, value in cells.items():
            try:
                row, col = map(int, key.split(','))
                ws.cell(row=row, column=col, value=value)
            except:
                print(f"⚠️ Помилка в ключі {key}")

    print(f"✅ Сейв завантажено: {json_path}")